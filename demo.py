#multi user working code -checking accurate results

class CaptchaDetectedException(Exception):
    def __init__(self, message, page_num=0):
        super().__init__(message)
        self.page_num = page_num

class TaskCancelledException(Exception):
    pass


import sys
sys.stdout.reconfigure(line_buffering=True) # Force unbuffered output for live logs

from flask import Flask, render_template_string, request, redirect, url_for, session, send_from_directory, send_file, jsonify
from flask_socketio import SocketIO, emit, join_room
import time, os, json, uuid, datetime, threading, argparse, random
import tempfile, shutil
import logging
import urllib.parse
# Database abstraction
import database as db_layer
from database import get_db_connection, init_db

# Reduce noisy logs
from dotenv import load_dotenv
load_dotenv()
logging.getLogger("engineio").setLevel(logging.ERROR)
logging.getLogger("socketio").setLevel(logging.ERROR)

from authlib.integrations.flask_client import OAuth

# ====================== CONFIG ======================
BASE_DIR = os.getcwd()

BRAVE_PATH = os.getenv("BRAVE_PATH", "")

EXTENSION_PATH = os.getenv(
    "EXTENSION_PATH",
    os.path.join(BASE_DIR, "extensions", "raptor_unpacked")
)

USERS_FILE = "users.json"
TOKENS_FILE = "tokens.json"
STATIC_FOLDER = "static"

# Shared tasks file (used to coordinate tasks between multiple processes).
TASKS_STORE = "rankplex_tasks.json"
print(f"DEBUG: TASKS_STORE set to {TASKS_STORE}")

# 32GB Server Optimization: Max 40 concurrent browsers
try:
    max_tasks = int(os.getenv("MAX_CONCURRENT_TASKS", "40"))
except ValueError:
    max_tasks = 40
task_semaphore = threading.Semaphore(max_tasks)
TASK_LOCK = threading.Lock() # Lock for task state transitions

# Email config
EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")

# Proxy config - Route browser traffic through server IP to reduce CAPTCHAs
PROXY_SERVER = os.getenv("PROXY_SERVER")
PROXY_PORT = os.getenv("PROXY_PORT", "8080")
PROXY_USERNAME = os.getenv("PROXY_USERNAME")
PROXY_PASSWORD = os.getenv("PROXY_PASSWORD")

# ====================== PROXY MANAGER ======================
class ProxyManager:
    """Manages a pool of static proxies or session persistence for rotating proxies."""
    def __init__(self, file_path="proxies.txt"):
        self.file_path = file_path
        self.static_proxies = []
        self.current_idx = 0
        self.sessions = {} # user_email -> session_id
        self.lock = threading.Lock()
        self.load_static_proxies()

    def load_static_proxies(self):
        if os.path.exists(self.file_path):
            try:
                with open(self.file_path, "r") as f:
                    self.static_proxies = [l.strip() for l in f if l.strip()]
                if self.static_proxies:
                    print(f"DEBUG: Loaded {len(self.static_proxies)} static proxies from {self.file_path}")
            except Exception as e:
                print(f"Error loading {self.file_path}: {e}")

    def get_static_proxy(self):
        with self.lock:
            if not self.static_proxies:
                return None
            proxy = self.static_proxies[self.current_idx]
            self.current_idx = (self.current_idx + 1) % len(self.static_proxies)
            return proxy

    def get_session(self, user_email):
        with self.lock:
            if user_email not in self.sessions:
                self.sessions[user_email] = str(uuid.uuid4().hex[:8])
            return self.sessions[user_email]

    def rotate_session(self, user_email):
        with self.lock:
            self.sessions[user_email] = str(uuid.uuid4().hex[:8])
            return self.sessions[user_email]

PROXY_MANAGER = ProxyManager()

# Geocoding cache: avoid repeated HTTP lookups when init_context() fires every 3 keywords
_geo_cache = {}  # location_str -> (geolocation_dict_or_None, country_code_or_None)


def parse_timestamp(ts):
    """Helper to convert SQLite string timestamp or datetime object to datetime object."""
    if not ts:
        return None
    if isinstance(ts, str):
        try:
            # SQLite default format: YYYY-MM-DD HH:MM:SS
            return datetime.datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            # Fallback or other formats
            try:
                return datetime.datetime.fromisoformat(ts)
            except ValueError:
                return None
    return ts

def add_notification(user_email, message):
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO notifications (user_email, message) VALUES (%s, %s)", (user_email, message))
            conn.commit()
            cursor.close()
            conn.close()
            # Emit live notification via socketio
            socketio.emit('new_notification', {'message': message}, room=user_email)
        except Exception as e:
            print(f"Error adding notification: {e}")

# Initialize database on startup


app = Flask(__name__, static_folder='.', static_url_path='')
app.secret_key = "super_secret_key_change_this_in_production_2025"

# Fix Google OAuth — use SameSite=Lax (browsers send Lax cookies on top-level GET redirects
# like OAuth callbacks). SameSite=None without Secure is rejected by all modern browsers.
app.config.update(
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=False,
    SESSION_COOKIE_HTTPONLY=True,
)

# Configure SERVER_NAME only for production to allow localhost testing
if os.getenv("FLASK_ENV") == "production":
    app.config.update(
        SERVER_NAME="rankplex.cloud",
        PREFERRED_URL_SCHEME="https",
        SESSION_COOKIE_SECURE=True,    # Use secure cookies in production (HTTPS)
    )


# THREADING backend (no eventlet/gevent/Redis)
socketio = SocketIO(
    app,
    cors_allowed_origins="*",
    async_mode="threading",
)

# Initialize OAuth
oauth = OAuth(app)
google = oauth.register(
    name="google",
    client_id = os.getenv("GOOGLE_CLIENT_ID"),
    client_secret = os.getenv("GOOGLE_CLIENT_SECRET"),
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_kwargs={"scope": "openid email profile"},
)


@app.route('/')
def index():
    return send_file('index.html')

@app.route('/about')
def about():
    return redirect('/#about')

@app.route('/features')
def features():
    return redirect('/#features')

@app.route('/how-it-works')
def how_it_works():
    return redirect('/#ranking-blk')

@app.route('/faq')
def faq():
    return redirect('/#faq')

@app.route('/payment')
def payment_page():
    return send_file('payment.html')

@app.route('/api/buy-credits', methods=['POST'])
def buy_credits():
    if 'user' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    data = request.json
    amount = data.get('amount')
    
    if not amount:
         return jsonify({'success': False, 'error': 'Invalid amount'}), 400
         
    user = session['user']
    # Handle if session['user'] is just the email string or a dict
    user_email = user['email'] if isinstance(user, dict) else user
    
    if db_layer.add_credits(user_email, int(amount)):
        return jsonify({'success': True, 'message': 'Credits added successfully'})
    else:
        return jsonify({'success': False, 'error': 'Database error'}), 500



@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'images'),
                               'fav.webp', mimetype='image/webp')

if not os.path.exists(STATIC_FOLDER):
    os.makedirs(STATIC_FOLDER)

# File-backed task storage
LOCK = threading.Lock()

# Worker ID from systemd (rankplex@N.service sets RANKPLEX_WORKER_ID=N)
WORKER_ID = os.environ.get("RANKPLEX_WORKER_ID", "1")
try:
    WORKER_ID_INT = int(WORKER_ID)
except ValueError:
    WORKER_ID_INT = 1




# ====================== TASKS FILE HELPERS ======================
def _load_tasks_from_disk():
    try:
        if os.path.exists(TASKS_STORE):
            with open(TASKS_STORE, "r") as f:
                return json.load(f)
    except Exception:
        return {}
    return {}


def _write_tasks_to_disk(tasks):
    tmp = TASKS_STORE + f".tmp.{uuid.uuid4().hex}"
    try:
        with open(tmp, "w") as f:
            json.dump(tasks, f)
        os.replace(tmp, TASKS_STORE)
    finally:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass


def get_task(run_id):
    with LOCK:
        tasks = _load_tasks_from_disk()
        return tasks.get(str(run_id))

def save_task(run_id, task):
    with LOCK:
        tasks = _load_tasks_from_disk()
        tasks[str(run_id)] = task
        _write_tasks_to_disk(tasks)

def delete_task_from_disk(run_id):
    with LOCK:
        tasks = _load_tasks_from_disk()
        if str(run_id) in tasks:
            del tasks[str(run_id)]
            _write_tasks_to_disk(tasks)

def get_active_tasks_for_user(user_email):
    active = []
    with LOCK:
        tasks = _load_tasks_from_disk()
        for _, t in tasks.items():
            if t.get('user') == user_email and t.get('status') in ["started", "waiting_for_socket", "ready_to_start"]:
                 if not t.get("cancelled"):
                     active.append(t)
    return active


def set_task_for_user(user, task_obj):
    with LOCK:
        tasks = _load_tasks_from_disk()
        tasks[user] = task_obj
        _write_tasks_to_disk(tasks)


def delete_task_for_user(user):
    with LOCK:
        tasks = _load_tasks_from_disk()
        if user in tasks:
            del tasks[user]
            _write_tasks_to_disk(tasks)


# ====================== USER & TOKEN HELPERS ======================
def load_users():
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, "r") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_users(users):
    with open(USERS_FILE, "w") as f:
        json.dump(users, f, indent=4)

def load_tokens():
    if os.path.exists(TOKENS_FILE):
        try:
            with open(TOKENS_FILE, "r") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_tokens(tokens):
    with open(TOKENS_FILE, "w") as f:
        json.dump(tokens, f, indent=4)


# ====================== EMAIL ======================
def send_verification_email(recipient, token):
    import smtplib
    from email.message import EmailMessage

    msg = EmailMessage()
    msg["Subject"] = "Verify Your Account - Nice Digitals"
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = recipient

    # Use request.host_url for dynamic domain
    link = f"{request.host_url}verify/{token}"
    msg.set_content(f"Welcome to Nice Digitals!\n\nPlease click the link below to verify your email address:\n\n{link}\n\nThis link is valid for 24 hours.")

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(msg)


def send_reset_email(recipient, token):
    import smtplib
    from email.message import EmailMessage

    msg = EmailMessage()
    msg["Subject"] = "Password Reset Request"
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = recipient

    link = f"{request.host_url}reset-password/{token}"
    msg.set_content(f"Click to reset password:\n\n{link}\n\nValid for 15 minutes.")

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(msg)


# ====================== PLAYWRIGHT HELPERS ======================
def wait_for_captcha_to_clear(page):
    """
    Simplified CAPTCHA detection. 
    With proxies, we rely more on the extension and IP rotation.
    """
    try:
        content_lower = page.content().lower()
        title_lower = page.title().lower()
        
        is_blocked = False
        if "unusual traffic" in content_lower or "verify you are human" in title_lower or "sorry" in title_lower:
             if "google" in title_lower or "google" in content_lower:
                 is_blocked = True
        
        if not is_blocked:
            for frame in page.frames:
                if frame.locator('.recaptcha-checkbox-border').count() > 0:
                    is_blocked = True
                    break
                    
        if is_blocked:
            print("DEBUG: Google CAPTCHA/Block detected. Signaling for rotation.")
            # We raise the exception to trigger worker-level rotation immediately.
            raise CaptchaDetectedException("Google CAPTCHA detected")
            
    except CaptchaDetectedException:
        raise
    except Exception as e:
        pass


def generate_uule(location_name):
    import base64
    
    # Simple UULE generator
    # Secret key for length
    secret = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-_"
    
    if not location_name:
        return ""

    try:
        # 1. Canonical Name must be valid
        # In a real app, we might need to look up the canonical name from Google's Adwords Geo targets.
        # Here we assume the user provides a valid location string like "New York,New York,United States".
        
        # 2. Get length character
        l = len(location_name)
        if l >= len(secret):
            return "" # Too long
            
        length_char = secret[l]
        
        # 3. Base64 encode the location
        # Encode string to bytes, then b64encode, then decode back to string
        b64_location = base64.b64encode(location_name.encode()).decode()
        
        # 4. Construct UULE
        uule = f"w+CAIQICI{length_char}{b64_location}"
        return uule
    except Exception:
        return ""





# Detailed Country Configuration for Emulation
COUNTRY_CONFIG = {
    "us": {"domain": "https://www.google.com/search", "timezone": "America/New_York", "locale": "en-US"},
    "uk": {"domain": "https://www.google.co.uk/search", "timezone": "Europe/London", "locale": "en-GB"},
    "gb": {"domain": "https://www.google.co.uk/search", "timezone": "Europe/London", "locale": "en-GB"},
    "au": {"domain": "https://www.google.com.au/search", "timezone": "Australia/Sydney", "locale": "en-AU"},
    "in": {"domain": "https://www.google.co.in/search", "timezone": "Asia/Kolkata", "locale": "en-IN"},
    "ca": {"domain": "https://www.google.ca/search", "timezone": "America/Toronto", "locale": "en-CA"},
    "de": {"domain": "https://www.google.de/search", "timezone": "Europe/Berlin", "locale": "de-DE"},
    "fr": {"domain": "https://www.google.fr/search", "timezone": "Europe/Paris", "locale": "fr-FR"},
    "jp": {"domain": "https://www.google.co.jp/search", "timezone": "Asia/Tokyo", "locale": "ja-JP"},
    "br": {"domain": "https://www.google.com.br/search", "timezone": "America/Sao_Paulo", "locale": "pt-BR"},
    "id": {"domain": "https://www.google.co.id/search", "timezone": "Asia/Jakarta", "locale": "id-ID"},
    "it": {"domain": "https://www.google.it/search", "timezone": "Europe/Rome", "locale": "it-IT"},
    "mx": {"domain": "https://www.google.com.mx/search", "timezone": "America/Mexico_City", "locale": "es-MX"},
    "nl": {"domain": "https://www.google.nl/search", "timezone": "Europe/Amsterdam", "locale": "nl-NL"},
    "es": {"domain": "https://www.google.es/search", "timezone": "Europe/Madrid", "locale": "es-ES"},
    "tr": {"domain": "https://www.google.com.tr/search", "timezone": "Europe/Istanbul", "locale": "tr-TR"},
    "se": {"domain": "https://www.google.se/search", "timezone": "Europe/Stockholm", "locale": "sv-SE"},
    "ae": {"domain": "https://www.google.ae/search", "timezone": "Asia/Dubai", "locale": "ar-AE"},
    "sa": {"domain": "https://www.google.com.sa/search", "timezone": "Asia/Riyadh", "locale": "ar-SA"},
    "sg": {"domain": "https://www.google.com.sg/search", "timezone": "Asia/Singapore", "locale": "en-SG"},
}

def get_coordinates_and_country(location_name):
    import requests
    if not location_name or location_name in ["Global", ""]:
        return None, None

    # ── Fast-path: return cached result to avoid repeated HTTP lookups per rotation ──
    if location_name in _geo_cache:
        print(f"DEBUG: Geocoding cache hit for '{location_name}'")
        return _geo_cache[location_name]
    
    # Simple retry loop
    for attempt in range(3):
        try:
            url = "https://nominatim.openstreetmap.org/search"
            params = {
                'q': location_name,
                'format': 'json',
                'limit': 1,
                'addressdetails': 1
            }
            headers = {
                'User-Agent': 'RankPlex/1.0'
            }
            resp = requests.get(url, params=params, headers=headers, timeout=5)
            data = resp.json()
            if data and len(data) > 0:
                item = data[0]
                lat = float(item['lat'])
                lon = float(item['lon'])
                country_code = item.get('address', {}).get('country_code', 'us')
                result = ({
                    'latitude': lat,
                    'longitude': lon,
                    'accuracy': 100
                }, country_code)
                _geo_cache[location_name] = result
                return result
        except Exception as e:
            print(f"Error Geocoding {location_name} (Attempt {attempt+1}): {e}")
            time.sleep(1)

    _geo_cache[location_name] = (None, None)  # Cache negative result too
    return None, None





def find_keyword_rank(page, keyword, target_domain, max_pages=10, location=None, country_code=None, start_page=0):
    domain_only = (
        target_domain.replace("https://", "")
        .replace("http://", "")
        .replace("www.", "")
        .strip("/")
    )
    
    
    # Base URL defaults
    base_url = "https://www.google.com/search"
    if country_code and country_code.lower() in COUNTRY_CONFIG:
        base_url = COUNTRY_CONFIG[country_code.lower()]['domain']
        
    uule = ""
    gl_param = ""
    
    if location and location not in ["Global", ""]:
        # If specific location provided
        generated_uule = generate_uule(location)
        if generated_uule:
            uule = f"&uule={generated_uule}"
        
        # Enforce strict 'near' location constraint
        encoded_location = urllib.parse.quote(location)
        uule += f"&near={encoded_location}"
            
    if country_code:
        gl_param = f"&gl={country_code}"
    
    # If no location or Global, we default to generic google.com. 
    # Current behavior was google.co.in. Let's stick to google.com for broader 'Global' sense,
    # or if the user wants India specifically they can type "India".
    # BUT, to respect previous default behavior if someone relied on it:
    # If location is explicitly "India" or None/Global, what should we do?
    # The requirement says: "when user enters location... browser location must be set"
    # So if they entered something, use it.
    
    # Loop through pages (10 results each) up to max_pages
    results_per_page = 10
    

    for p in range(start_page, max_pages):
        start = p * results_per_page
        encoded_keyword = urllib.parse.quote(keyword)
        url = f"{base_url}?q={encoded_keyword}&hl=en&start={start}{uule}{gl_param}&filter=0&ie=UTF-8"
        
        print(f"DEBUG: [Keyword: {keyword}] Checking Page {p+1} (start={start}): {url}")
        
        try:
            # 1. Navigate and wait for full page content including JS-rendered results
            page.goto(url, wait_until="domcontentloaded", timeout=60000)
            
            # Give JS time to render after domcontentloaded (reduced: 1.5s → 0.8s)
            time.sleep(0.8)

            # Wait for organic results OR detect a block/error page
            results_appeared = False
            try:
                page.wait_for_selector("#search a:has(h3), #rso a:has(h3)", timeout=4000)  # 7s→4s — pages load in 2-3s
                results_appeared = True
            except:
                # Selector not found — could be CAPTCHA, no results, or unusual layout
                # We'll handle this in the CAPTCHA check below
                print(f"DEBUG: #{p+1} selector timeout for '{keyword}' — may be CAPTCHA or no results")

            # 2. Handle Google Consent/Cookie popups
            try:
                for btn_text in ["Accept all", "I agree", "Accept everything", "Agree"]:
                    btn = page.locator(f'button:has-text("{btn_text}")').first
                    if btn.is_visible(timeout=500):  # 1s→500ms — consent popup is instant
                        print(f"DEBUG: Google Consent popup — clicking '{btn_text}'")
                        btn.click()
                        page.wait_for_load_state("domcontentloaded")
                        time.sleep(1)
                        # After consent, wait again for results
                        try:
                            page.wait_for_selector("#search a:has(h3), #rso a:has(h3)", timeout=7000)
                            results_appeared = True
                        except:
                            pass
                        break
            except:
                pass


            # 3. Check for CAPTCHA / block pages
            page_html = page.content().lower()
            page_title = page.title().lower()
            if ("unusual traffic" in page_html or "verify you are human" in page_html 
                    or "sorry..." in page_title or "about this page" in page_html 
                    or "recaptcha" in page_html):
                print(f"DEBUG: [Keyword: {keyword}] CAPTCHA/Block detected at Page {p+1}.")
                raise CaptchaDetectedException("Google block detected", p)

            # 4. Use JavaScript to extract organic links — reliable across layout changes
            # The key insight: Google's ad blocks sit in #tads (top) and #bottomads (bottom).
            # Everything inside #rso (the main results div) is organic.
            extract_script = """
            () => {
                const organic = [];
                const seen = new Set();

                // #rso = Real Search Results Only — the clean organic list
                // This is more precise than #search which includes ads
                const mainResults = document.querySelector('#rso');
                const searchDiv = document.querySelector('#search');
                const container = mainResults || searchDiv;

                if (!container) return organic;

                // Collect ALL ad/sponsored block nodes to skip them later
                const adIds = ['tads', 'bottomads', 'tadsb', 'commercial-unit-desktop-top', 'bres'];
                const adNodes = [];
                adIds.forEach(id => {
                    const el = document.getElementById(id);
                    if (el) adNodes.push(el);
                });
                // Also collect by class
                document.querySelectorAll('[data-text-ad], .commercial-unit-desktop-top, .cu-container').forEach(el => adNodes.push(el));

                function isInsideAd(node) {
                    return adNodes.some(ad => ad.contains(node));
                }

                // Find all links that contain an H3 heading (= result titles)
                const allLinks = container.querySelectorAll('a:has(h3)');

                allLinks.forEach(link => {
                    if (isInsideAd(link)) return; // Skip sponsored

                    const href = link.href;
                    if (!href || href.startsWith('#') || href.includes('javascript:')) return;
                    // Skip Google's own internal navigation links
                    if (href.includes('google.') && !href.includes('/url?')) return;

                    // Skip People Also Ask, Videos, Images sections
                    const parentText = (link.closest('div[jscontroller], div.g, div[data-hveid]') || link.parentElement || link);
                    const blockLabel = parentText.innerText ? parentText.innerText.substring(0, 40).toLowerCase() : '';
                    if (blockLabel.startsWith('people also ask') || blockLabel.startsWith('videos') || blockLabel.startsWith('images for')) return;

                    if (!seen.has(href)) {
                        seen.add(href);
                        organic.push(href);
                    }
                });

                return organic;
            }
            """
            found_on_page = page.evaluate(extract_script)
            found_on_page = [h for h in (found_on_page or []) if h]  # Remove None/empty

            print(f"DEBUG: [Keyword: {keyword}] Page {p+1}: JS extracted {len(found_on_page)} organic candidates.")
            if found_on_page:
                print(f"DEBUG:   First 5: {found_on_page[:5]}")

            if not found_on_page:
                # Genuine no-results check
                if "did not match any documents" in page_html or "no results found" in page_html:
                    print(f"DEBUG: Verified no results for '{keyword}' on page {p+1}")
                    return None, None, None
                # Probably still a block or JS didn't run
                print(f"DEBUG: Zero links on Page {p+1} for '{keyword}' — treating as block.")
                raise CaptchaDetectedException("Zero organic links extracted", p)

            # 5. MATCHING: Find target domain in extracted organic links
            target = domain_only.lower().strip("/")
            for i, href in enumerate(found_on_page):
                # Resolve Google redirect URLs like /url?q=https://...
                raw_href = href
                if "/url?" in href:
                    try:
                        qs = urllib.parse.parse_qs(urllib.parse.urlparse(href).query)
                        raw_href = qs.get("q", qs.get("url", [href]))[0]
                    except:
                        pass

                # Normalize: remove scheme, www
                res_url = raw_href.lower().replace("https://", "").replace("http://", "").replace("www.", "").strip("/")

                # Match: target 'flipkart.com' should match 'flipkart.com/mobiles', 'm.flipkart.com/...', etc.
                is_match = (
                    res_url == target
                    or res_url.startswith(target + "/")
                    or res_url.startswith(target + "?")
                    or ("." + target) in res_url  # subdomains like m.flipkart.com
                )

                if is_match:
                    rank_overall = start + i + 1
                    print(f"DEBUG: [SUCCESS] '{keyword}' -> {target} at Page {p+1}, Overall Rank {rank_overall} ({raw_href[:80]})")
                    return rank_overall, p + 1, raw_href

            print(f"DEBUG: Page {p+1} done for '{keyword}'. Target not found, continuing...")
            time.sleep(random.uniform(0.5, 1.5))  # Reduced page-gap delay: 1-3s → 0.5-1.5s

        except CaptchaDetectedException:
            raise  # Worker handles rotation
        except Exception as e:
            err_msg = str(e)
            print(f"DEBUG: Error on Page {p+1} for '{keyword}': {err_msg}")
            if "timeout" in err_msg.lower() or "closed" in err_msg.lower() or "crashed" in err_msg.lower():
                raise CaptchaDetectedException(f"Browser/network error: {err_msg}", p)
            continue  # Try next page for non-critical errors

    return None, None, None




# ====================== EXCEL WRITER ======================
def write_results_with_layout(file_path, target_domain, df):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    df_to_write = df[["Keyword", "Page", "Rank"]].copy()
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df_to_write.to_excel(writer, index=False, startrow=5, header=False)

    wb = load_workbook(file_path)
    ws = wb.active
    ws["A1"] = "Target Domain"
    ws["B1"] = target_domain
    ws["A2"] = "Search Engine"
    ws["B2"] = "Google"
    ws["A3"] = "Date"
    ws["B3"] = datetime.datetime.now().strftime("%d-%m-%Y, %H:%M")
    ws["A5"], ws["B5"], ws["C5"] = "Keyword", "Page", "Rank"
    bold = Font(bold=True)
    for cell in ["A1", "A2", "A3", "A5", "B5", "C5"]:
        ws[cell].font = bold
    wb.save(file_path)



# ====================== PLAYWRIGHT BROWSER ======================

# Modern User Agents List (Windows-focused for consistency)
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:132.0) Gecko/20100101 Firefox/132.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
]

def launch_worker_browser(extension_path=None):
    from playwright.sync_api import sync_playwright
    import random

    p = sync_playwright().start()

    # Use provided path or fallback to global
    ext_to_use = extension_path if extension_path else EXTENSION_PATH

    # Validate Extension Path
    if not ext_to_use or not os.path.exists(ext_to_use):
        print(f"CRITICAL ERROR: Extension path not found at {ext_to_use}")
        # We will continue WITHOUT extension to avoid hang, but warn user
        args = [
            "--disable-dev-shm-usage",
            "--disable-gpu",
            "--disable-blink-features=AutomationControlled",
        ]
    else:
        # ULTRA-LIGHT MODE: Extensions disabled to save RAM since we rotate on detection
        args = [
            "--disable-dev-shm-usage",
            "--disable-blink-features=AutomationControlled",
            "--no-first-run",
            "--no-default-browser-check",
            "--password-store=basic",
        ]
        
    print(f"DEBUG: OS Name detected: {os.name}")

    args.extend([
        # WEBRTC & PRIVACY
        "--disable-webrtc", 
        "--disable-webrtc-multiple-routes",
        "--enforce-webrtc-ip-permission-check",
        "--force-webrtc-ip-handling-policy=disable_non_proxied_udp",
        
        # ADDITIONAL STEALTH & BANDWIDTH SAVING
        "--disable-blink-features=AutomationControlled", 
        "--disable-infobars",
        "--exclude-switches=enable-automation",
        "--use-fake-ui-for-media-stream",
        "--lang=en-US,en;q=0.9",
        
        # BANDWIDTH OPTIMIZATIONS (Disable background noise)
        "--disable-background-networking",
        "--disable-background-timer-throttling",
        "--disable-backgrounding-occluded-windows",
        "--disable-breakpad",
        "--disable-client-side-phishing-detection",
        "--disable-component-update",
        "--disable-default-apps",
        "--disable-domain-reliability",
        "--disable-extensions",
        "--disable-features=AudioServiceOutOfProcess,IsolateOrigins,site-per-process",
        "--disable-hang-monitor",
        "--disable-ipc-flooding-protection",
        "--disable-notifications",
        "--disable-offer-store-unmasked-wallet-cards",
        "--disable-popup-blocking",
        "--disable-print-preview",
        "--disable-prompt-on-repost",
        "--disable-renderer-backgrounding",
        "--disable-speech-api",
        "--disable-sync",
        "--hide-scrollbars",
        "--metrics-recording-only",
        "--mute-audio",
        "--no-pings",
        "--use-mock-keychain",
    ])

    # Only add sandbox disabling flags on Linux/Headless where absolutely needed
    if os.name == 'posix': # Linux/Unix
         args.append("--disable-setuid-sandbox")

    launch_kwargs = {
        "headless": False, # HIGH SPEED & LOW RAM
        "args": args,
        "ignore_default_args": ["--enable-automation", "--enable-blink-features=IdleDetection"],
    }
        
    if BRAVE_PATH and os.path.exists(BRAVE_PATH):
        launch_kwargs["executable_path"] = BRAVE_PATH

    print(f"DEBUG: LAUNCHING BROWSER ENGINE")
    browser = p.chromium.launch(**launch_kwargs)

    return p, browser

def create_browser_context(browser, geolocation=None, timezone_id="UTC", locale="en-US", target_country=None):
    import random
    import uuid

    # Random Viewport
    viewports = [
        {"width": 1920, "height": 1080},
        {"width": 1366, "height": 768},
        {"width": 1536, "height": 864},
        {"width": 1440, "height": 900},
        {"width": 1280, "height": 720},
    ]
    viewport = random.choice(viewports)
    
    # Custom User Agent
    user_agent = random.choice(USER_AGENTS)

    context_kwargs = {
        "accept_downloads": False,
        "geolocation": geolocation,
        "permissions": ["geolocation"] if geolocation else [],
        "timezone_id": timezone_id,
        "locale": locale,
        "viewport": viewport,
        "user_agent": user_agent,
        "ignore_https_errors": True, # Ensure proxy certs don't block
        "device_scale_factor": 1,
    }

    # Proxy Configuration for this Context
    static_proxy = PROXY_MANAGER.get_static_proxy()
    if static_proxy:
        if "@" in static_proxy:
            context_kwargs["proxy"] = {"server": static_proxy}
        elif static_proxy.count(":") == 3:
            h, p_port, u, pw = static_proxy.split(":")
            context_kwargs["proxy"] = {
                "server": f"http://{h}:{p_port}",
                "username": u,
                "password": pw
            }
        else:
            context_kwargs["proxy"] = {"server": f"http://{static_proxy}"}
        print(f"DEBUG: Using static proxy: {static_proxy}")
        
    elif PROXY_SERVER:
        proxy_url = f"http://{PROXY_SERVER}:{PROXY_PORT}"
        proxy_config = {"server": proxy_url}
        if PROXY_USERNAME and PROXY_PASSWORD:
            proxy_config["username"] = PROXY_USERNAME
            proxy_config["password"] = PROXY_PASSWORD
        context_kwargs["proxy"] = proxy_config
        print(f"DEBUG: Context Proxy Session created: {proxy_config.get('username', 'anonymous')}")

    context = browser.new_context(**context_kwargs)
    page = context.new_page()

    try:
        page.add_init_script(
            """
            Object.defineProperty(navigator, 'webdriver', {get: () => false});
            window.navigator.chrome = { runtime: {}, };
            Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en']});
            Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
            Object.defineProperty(navigator, 'deviceMemory', {get: () => 8});
            Object.defineProperty(navigator, 'hardwareConcurrency', {get: () => 8});
            
            // WebGL detection masking
            const getParameter = WebGLRenderingContext.prototype.getParameter;
            WebGLRenderingContext.prototype.getParameter = function(parameter) {
                if (parameter === 37445) return 'Intel Open Source Technology Center';
                if (parameter === 37446) return 'Mesa DRI Intel(R) Ivybridge Mobile ';
                return getParameter.apply(this, arguments);
            };
            """
        )
    except Exception:
        pass

    return context, page


# ====================== BACKGROUND TASK ======================
def run_rank_checker(keywords, target_domain, max_pages, user_email, sid, project_name="No Project", location="Global", run_id=None):
    print(f"DEBUG: Entering run_rank_checker for run_id {run_id} (Keyword count: {len(keywords)})")
    import pandas as pd
    import concurrent.futures
    import queue
    import traceback
    import gc # Memory management optimization

    # Non-blocking acquire; if no slot, return
    if not task_semaphore.acquire(blocking=False):
        socketio.emit(
            "status_update",
            {"error": "Server busy (max tasks). Try again later."},
            room=f"task_{run_id}",
        )
        return

    # Initialize shared state
    progress_lock = threading.Lock()
    processed_count = 0
    total = len(keywords)
    combined_data_rows = []
    
    # Use a Queue for dynamic load balancing
    keyword_queue = queue.Queue()
    for i, k in enumerate(keywords):
        keyword_queue.put((i, k))

    # Save run metadata to database if new run
    conn = get_db_connection()
    if not run_id and conn:
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO runs (user_email, project_name, target_domain, location, max_pages, total_keywords)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (user_email, project_name, target_domain, location, max_pages, len(keywords)))
            run_id = cursor.lastrowid
            conn.commit()
            cursor.close()
        except Exception as err:
            print(f"Error saving run to DB: {err}")
    if conn:
        conn.close()

    add_notification(user_email, f"Started checking with {target_domain}")
    
    try:
        socketio.emit(
            "status_update",
            {"msg": "Connected to Google, utilizing cloud queue...", "run_id": run_id},
            room=f"task_{run_id}",
        )

        def worker_loop(worker_idx):
            print(f"DEBUG: Worker {worker_idx} started.")
            nonlocal processed_count
            worker_results = []
            
            # Counter for memory management restart
            keywords_processed_in_session = 0
            
            p_worker = None
            browser_worker = None
            context_worker = None
            page_worker = None
            conn_worker = None
            worker_country_code = None
            ip_captcha_count = 0
            
            def init_browser():
                nonlocal p_worker, browser_worker
                try:
                    if browser_worker:
                        try: browser_worker.close()
                        except: pass
                    if p_worker:
                        try: p_worker.stop()
                        except: pass
                except:
                    pass
                
                # Use the unique extension copy if available (DISABLED for RAM optimization)
                ext_path_arg = None
                
                p_worker, browser_worker = launch_worker_browser(extension_path=ext_path_arg)
                
            def init_context():
                nonlocal context_worker, page_worker, worker_country_code, browser_worker, keywords_processed_in_session
                try:
                    # More aggressive cleanup to ensure old tabs are removed
                    if page_worker:
                        try: page_worker.close()
                        except: pass
                    if context_worker:
                        try: context_worker.close()
                        except: pass
                except:
                    pass
                
                geolocation, country_code = get_coordinates_and_country(location)
                worker_country_code = country_code # Store for find_keyword_rank
                timezone_id = "UTC"
                locale = "en-US"
                if country_code and country_code.lower() in COUNTRY_CONFIG:
                    cfg = COUNTRY_CONFIG[country_code.lower()]
                    timezone_id = cfg.get("timezone", "UTC")
                    locale = cfg.get("locale", "en-US")
                
                context_worker, page_worker = create_browser_context(
                    browser_worker, geolocation, timezone_id, locale,
                    target_country=country_code  # Pass country so proxy uses correct geo IP
                )
                
                # Reset counter to guarantee rotation after certain ops
                keywords_processed_in_session = 0
                ip_captcha_count = 0

                # NON-BLOCKING background IP logger: logs proxy IP without stalling keyword processing
                def _log_proxy_ip(w_idx):
                    try:
                        import requests as _req
                        resp = _req.get("http://api.myip.com", timeout=8)
                        ip_data = resp.json()
                        print(f"[Worker {w_idx}] Proxy IP: {ip_data.get('ip')} | Country: {ip_data.get('country')} ({ip_data.get('cc')})")
                    except Exception:
                        print(f"[Worker {w_idx}] Could not fetch proxy IP (non-critical)")
                
                # BANDWIDTH ROUTING: Block images, CSS, fonts, media — but ALLOW scripts and XHR
                try:
                    def should_block(request):
                        resource_type = request.resource_type
                        url = request.url.lower()

                        if resource_type in ["image", "stylesheet", "font", "media", "manifest"]:
                            return True

                        ad_block_patterns = [
                            "google-analytics.com", "googletagmanager.com",
                            "googleadservices.com", "doubleclick.net",
                            "facebook.com", "facebook.net", "ytimg.com",
                            "adservice.google",
                        ]
                        if any(p in url for p in ad_block_patterns):
                            return True

                        return False

                    context_worker.route("**/*", lambda route: route.abort() if should_block(route.request) else route.continue_())
                except Exception as e_route:
                    print(f"Routing setup error: {e_route}")

            try:
                # Initial browser launch
                init_browser()
                init_context()
                conn_worker = get_db_connection()

                while True:
                    try:
                        # Non-blocking get with timeout to exit if queue empty
                        idx, k = keyword_queue.get(block=False)
                    except queue.Empty:
                        break

                    # Desynchronize workers slightly at each keyword to avoid "burst" patterns (reduced: 1-4s → 0.3-1.5s)
                    time.sleep(random.uniform(0.3, 1.5))
                    
                    # Check cancellation
                    current_task = get_task(run_id)
                    if current_task is None or current_task.get("cancelled"):
                        break

                    try:
                        # UNLIMITED RETRY LOOP: We rotate proxies until we get a result. 
                        # This fulfills the requirement: "it has to check whatever the way"
                        max_keyword_retries = 9999 
                        keyword_retry_count = 0
                        rank, page_num, landing_url = None, None, None
                        current_start_page = 0
                        tabs_to_close = []
                        
                        while True:
                            try:
                                # Process delayed tab closures securely within the main Playwright thread
                                current_time = time.time()
                                for tb, close_time in list(tabs_to_close):
                                    if current_time >= close_time:
                                        try: tb.close()
                                        except: pass
                                        tabs_to_close.remove((tb, close_time))
                                
                                # Check cancellation inside retry loop so we don't get stuck forever
                                current_task_check = get_task(run_id)
                                if current_task_check is None or current_task_check.get("cancelled"):
                                    raise TaskCancelledException("Task was cancelled by user")
                                    
                                # Browser/Page health check
                                if page_worker is None or page_worker.is_closed():
                                    init_browser()
                                
                                # PROXY HEALTH CHECK: Removed from hot path — no longer blocking keyword processing.
                                # IP logging now happens in a background thread inside init_context() on each rotation.
                                
                                # Attempt search
                                rank, page_num, landing_url = find_keyword_rank(page_worker, k, target_domain, max_pages, location, worker_country_code, current_start_page)
                                ip_captcha_count = 0  # Reset Captcha count on success so the IP is preserved
                                break # Success: exit loop
                                
                            except CaptchaDetectedException as e_cap:
                                current_start_page = e_cap.page_num
                                keyword_retry_count += 1
                                if keyword_retry_count >= 50:
                                    print(f"[Worker {worker_idx}] Max 50 proxy rotations reached for '{k}'. Moving on.")
                                    break
                                
                                ip_captcha_count += 1
                                if ip_captcha_count > 2:  # Changed from >3 → rotate IP faster after 2 CAPTCHAs
                                    print(f"[Worker {worker_idx}] CAPTCHA/Block > 2 times for this IP. Changing IP fast...")
                                    init_context()
                                else:
                                    print(f"[Worker {worker_idx}] CAPTCHA/Block detected for '{k}' resuming from page {current_start_page} (Attempt {keyword_retry_count}). Opening new tab in same browser, keeping IP...")
                                    new_page = context_worker.new_page()
                                    
                                    # Append the old page to be closed in the next iterations securely
                                    tabs_to_close.append((page_worker, time.time() + 5))
                                    
                                    page_worker = new_page
                                    
                                # Capped backoff: reduced from max 30s → max 8s to unblock faster
                                sleep_time = min(keyword_retry_count * 1, 8)
                                # Still check cancellation during wait
                                for _ in range(sleep_time):
                                    chk = get_task(run_id)
                                    if chk is None or chk.get("cancelled"):
                                        raise TaskCancelledException("Task was cancelled by user")
                                    time.sleep(1)
                            except TaskCancelledException:
                                print(f"[Worker {worker_idx}] Task stopped forcefully. Exiting worker.")
                                raise # Bubble up directly to exit keyword processing
                            except Exception as e_search:
                                keyword_retry_count += 1
                                if keyword_retry_count >= 50:
                                    print(f"[Worker {worker_idx}] Max retries reached for '{k}': {e_search}")
                                    break
                                print(f"[Worker {worker_idx}] Search error for '{k}' returning from page {current_start_page}: {e_search}. Retrying...")
                                try:
                                    new_page = context_worker.new_page()
                                    tabs_to_close.append((page_worker, time.time() + 2))
                                    page_worker = new_page
                                except Exception:
                                    init_context()
                                time.sleep(1)
                                chk = get_task(run_id)
                                if chk is None or chk.get("cancelled"):
                                    raise TaskCancelledException("Task was cancelled by user")
                        
                        if keyword_retry_count >= max_keyword_retries:
                            # This block is theoretically unreachable now but kept for structural integrity
                            print(f"[Worker {worker_idx}] Giving up on '{k}' after {max_keyword_retries} proxy rotations.")
                            rank, page_num, landing_url = None, None, None

                        # PROACTIVE ROTATION LOGIC: Rotate IP after exactly 3 keywords
                        keywords_processed_in_session += 1
                        if keywords_processed_in_session >= 3:
                            print(f"[Worker {worker_idx}] Preemptive rotation triggered ({keywords_processed_in_session} keywords). Switching proxy IP.")
                            init_context()

                        display_rank = str(rank) if rank else "Not found in top 100"
                        display_page = str(page_num) if page_num else ""
                        
                        row_data = {
                            "Keyword": k,
                            "Page": display_page,
                            "Rank": display_rank,
                            "Landing Page": landing_url or "",
                            "sort_order": idx
                        }
                        worker_results.append(row_data)

                        with progress_lock:
                            processed_count += 1
                            current_progress_val = int(processed_count * 100 / total)
                            socketio.emit(
                                "status_update",
                                {
                                    "msg": f"Checked: {k} ( {processed_count} of {total} )",
                                    "progress": current_progress_val,
                                },
                                room=f"task_{run_id}", 
                            )
                            if processed_count % 5 == 0:
                                if current_task:
                                    current_task["processed_count"] = processed_count
                                    current_task["last_update"] = datetime.datetime.now().isoformat()
                                    save_task(run_id, current_task)

                        # Individual DB Save
                        if conn_worker and run_id:
                            try:
                                cursor = conn_worker.cursor()
                                cursor.execute("""
                                    INSERT INTO results (run_id, keyword, page, rank, landing_page, sort_order)
                                    VALUES (%s, %s, %s, %s, %s, %s)
                                """, (run_id, k, display_page, display_rank, landing_url or "", idx))
                                conn_worker.commit()
                                cursor.close()
                            except Exception as err:
                                print(f"Error saving result to DB: {err}")
                        
                        # Individual DB Save

                    except Exception as e_inner:
                        print(f"[Worker {worker_idx}] Error processing keyword '{k}': {e_inner}")
                        # Prevent rapid-fire failures if browser is broken
                        time.sleep(2)
                        traceback.print_exc()
                        # Allow continue to next keyword
                    finally:
                        keyword_queue.task_done()

            except Exception as e_outer:
                print(f"[Worker {worker_idx}] Fatal Loop Error: {e_outer}")
            finally:
                try:
                    if context_worker: context_worker.close()
                    if browser_worker: browser_worker.close()
                    if p_worker: p_worker.stop()
                    if conn_worker: conn_worker.close()
                except:
                    pass
                try:
                    if os.path.exists(worker_ext_dir):
                        shutil.rmtree(worker_ext_dir, ignore_errors=True)
                except:
                    pass
            
            print(f"DEBUG: Worker {worker_idx} finished processing queue. Closing browser.")
            return worker_results

        # Start Parallel Workers (Optimized for 32GB Server)
        num_workers = 3
        print(f"DEBUG: Starting {num_workers} queue-based workers")
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=num_workers) as executor:
            futures = [executor.submit(worker_loop, i) for i in range(num_workers)]
            
            for f in concurrent.futures.as_completed(futures):
                try:
                    res = f.result()
                    if res:
                        combined_data_rows.extend(res)
                except Exception as e:
                    print(f"Worker execution failed: {e}")
        
        # Sort combined_data_rows by sort_order to maintain original input order
        combined_data_rows.sort(key=lambda x: x.get('sort_order', 0))

        # Finalize
        current_task_check = get_task(run_id)
        if current_task_check and current_task_check.get("cancelled"):
            socketio.emit("status_update", {"error": "Task cancelled."}, room=f"task_{run_id}")
            return

        if not combined_data_rows:
            # If no data collected, create empty DF with expected columns to avoid KeyError
            df = pd.DataFrame(columns=["Keyword", "Page", "Rank", "Landing Page"])
        else:
            df = pd.DataFrame(combined_data_rows)
        
        clean_domain_name = target_domain.replace("https://", "").replace("http://", "").replace("www.", "").strip("/")
        if "." in clean_domain_name:
             clean_domain_name = clean_domain_name.split(".")[0]
        clean_domain_name = "".join([c for c in clean_domain_name if c.isalnum() or c in ('-', '_')])
        if not clean_domain_name:
             clean_domain_name = "rank_check"
             
        current_year = datetime.datetime.now().year
        filename = f"{clean_domain_name}_{current_year}.xlsx"
        path = os.path.join(STATIC_FOLDER, filename)
        
        write_results_with_layout(path, target_domain, df)

        with open(path, 'rb') as f:
            excel_binary = f.read()

        conn_final = get_db_connection()
        if conn_final and run_id:
            try:
                cursor = conn_final.cursor()
                cursor.execute("UPDATE runs SET excel_filename = %s, excel_data = %s WHERE id = %s", (filename, excel_binary, run_id))
                conn_final.commit()
                cursor.close()
                conn_final.close()
            except Exception as err:
                print(f"Error updating run with filename and data: {err}")

        socketio.emit(
            "status_update",
            {
                "done": True,
                "msg": "Completed! Download ready.",
                "filename": filename,
                "run_id": run_id,
                "progress": 100,
            },
            room=f"task_{run_id}",
        )
        add_notification(user_email, f"Completed checking for {target_domain}")

        current_task = get_task(run_id)
        if current_task:
            current_task["status"] = "completed"
            current_task["run_id"] = run_id
            current_task["filename"] = filename
            save_task(run_id, current_task)

    except Exception as e:
        print(f"Global task error: {e}")
        current_task = get_task(run_id)
        if current_task:
            current_task["status"] = "failed"
            current_task["error"] = str(e)
            save_task(run_id, current_task)
        socketio.emit("status_update", {"error": f"Task failed: {str(e)}"}, room=f"task_{run_id}")

    finally:
        try:
            delete_task_for_user(user_email)
        except:
            pass
        task_semaphore.release()


# ====================== ROUTES ======================

@app.route("/health")
def health():
    return "OK", 200


@app.route("/api/notifications")
def get_notifications():
    user = session.get("user")
    if not user:
        return jsonify([])
    
    notifications = []
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM notifications WHERE user_email = %s ORDER BY timestamp DESC", (user,))
            rows = cursor.fetchall()
            notifications = [dict(row) for row in rows]
            # Format timestamp
            for n in notifications:
                dt = parse_timestamp(n['timestamp'])
                if dt:
                    n['timestamp'] = dt.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    n['timestamp'] = str(n['timestamp'])
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Error fetching notifications: {e}")
            
    return jsonify(notifications)

@app.route("/api/notifications/read", methods=['POST'])
def mark_notifications_read():
    user = session.get("user")
    if not user:
        return jsonify({"success": False})
    
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("UPDATE notifications SET is_read = 1 WHERE user_email = %s", (user,))
            conn.commit()
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Error marking notifications as read: {e}")
            
    return jsonify({"success": True})

@app.route("/api/notifications/delete-all", methods=['POST'])
def clear_notifications():
    user = session.get("user")
    if not user:
        return jsonify({"success": False})
    
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM notifications WHERE user_email = %s", (user,))
            conn.commit()
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Error clearing notifications: {e}")
            
    return jsonify({"success": True})



@app.route("/api/delete-run/<int:run_id>", methods=['POST', 'DELETE'])
def delete_run(run_id):
    user = session.get("user")
    if not user:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            # Verify ownership
            cursor.execute("SELECT user_email FROM runs WHERE id = %s", (run_id,))
            run = cursor.fetchone()
            if not run or run[0] != user:
                cursor.close()
                conn.close()
                return jsonify({"success": False, "error": "Unauthorized"}), 403
            
            # Delete results first (cascading normally but being safe)
            cursor.execute("DELETE FROM results WHERE run_id = %s", (run_id,))
            # Delete the run itself
            cursor.execute("DELETE FROM runs WHERE id = %s", (run_id,))
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"success": True})
        except Exception as e:
            print(f"Error deleting run {run_id}: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
            
    return jsonify({"success": False, "error": "Database connection failed"}), 500


@app.route("/download-result/<int:run_id>")
def download_result(run_id):
    user = session.get("user")
    if not user:
        return redirect(url_for("home"))

    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT project_name, target_domain, excel_filename, excel_data FROM runs WHERE id = %s AND user_email = %s", (run_id, user))
            run_data = cursor.fetchone()
            
            if not run_data:
                cursor.close()
                conn.close()
                return "Run not found or unauthorized", 404

            if run_data['excel_data']:
                cursor.close()
                conn.close()
                import io
                return send_file(
                    io.BytesIO(run_data['excel_data']),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=run_data['excel_filename'] or f"results_{run_id}.xlsx"
                )
            else:
                # No blob found (partial/stopped run), generate from results table
                cursor.execute("SELECT keyword, page, rank FROM results WHERE run_id = %s", (run_id,))
                results = cursor.fetchall()
                cursor.close()
                conn.close()

                if not results:
                     return "No results found for this run yet.", 404
                
                # Ensure results are dicts (fix for SQLite Row objects vs MySQL dicts)
                results = [dict(row) for row in results]

                # Generate Excel on the fly
                import pandas as pd
                import io
                from openpyxl import Workbook
                from openpyxl.styles import Font

                df = pd.DataFrame(results)
                # Ensure columns logic
                # Rename if needed to match export format
                df = df.rename(columns={"keyword": "Keyword", "page": "Page", "rank": "Rank"})
                
                output = io.BytesIO()
                
                # We can't use write_results_with_layout easily with BytesIO without some adaptation or temp file
                # But we can replicate the logic using openpyxl directly or pandas
                
                # Simple pandas export for now, or nice layout?
                # Let's try to match the layout
                wb = Workbook()
                ws = wb.active
                
                target_domain = run_data['target_domain']
                ws["A1"] = "Target Domain"
                ws["B1"] = target_domain
                ws["A2"] = "Search Engine"
                ws["B2"] = "Google"
                ws["A3"] = "Date"
                ws["B3"] = datetime.datetime.now().strftime("%d-%m-%Y, %H:%M")
                
                ws["A5"] = "Keyword"
                ws["B5"] = "Page"
                ws["C5"] = "Rank"
                
                bold = Font(bold=True)
                for cell in ["A1", "A2", "A3", "A5", "B5", "C5"]:
                    ws[cell].font = bold
                
                # Write data starting at row 6
                for idx, row in df.iterrows():
                    r = idx + 6
                    ws.cell(row=r, column=1, value=row['Keyword'])
                    ws.cell(row=r, column=2, value=row['Page'])
                    ws.cell(row=r, column=3, value=row['Rank'])

                wb.save(output)
                output.seek(0)
                
                filename = run_data['excel_filename'] or f"results_partial_{run_id}.xlsx"
                
                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )

        except Exception as e:
            print(f"Error downloading/generating from DB: {e}")
            import traceback
            traceback.print_exc()
            return f"Internal error: {str(e)}", 500
    return "DB Connection failed", 500


@app.route("/terms")
def terms_page():
    with open(os.path.join(os.path.dirname(__file__), 'terms-conditions.html'), 'r', encoding='utf-8') as f:
        content = f.read()
    return render_template_string(content)


@app.route("/privacy")
def privacy_page():
    with open(os.path.join(os.path.dirname(__file__), 'privacy-policy.html'), 'r', encoding='utf-8') as f:
        content = f.read()
    return render_template_string(content)


@app.route("/")


@app.route("/")
def home():
    if "user" in session:
        return redirect(url_for("dashboard"))
    # Updated to show index.html as the landing page
    return send_file(os.path.join(os.path.dirname(__file__), 'index.html'))


@app.route("/signup")
def signup_page():
    print("DEBUG: Hit signup route")
    with open(os.path.join(os.path.dirname(__file__), 'sign-up.html'), 'r', encoding='utf-8') as f:
        content = f.read()
    return render_template_string(content, error="")


@app.route("/signup", methods=["POST"])
def signup():
    email = request.form.get("email")
    password = request.form.get("password")
    fullname = request.form.get("fullname")
    
    users = load_users()
    if email in users:
        with open(os.path.join(os.path.dirname(__file__), 'sign-up.html'), 'r', encoding='utf-8') as f:
            content = f.read()
        return render_template_string(content, error="User already exists! Please login.")
    
    is_valid, msg = validate_password_strength(password)
    if not is_valid:
        with open(os.path.join(os.path.dirname(__file__), 'sign-up.html'), 'r', encoding='utf-8') as f:
            content = f.read()
        return render_template_string(content, error=msg)
    
    # Save credential (inactive)
    users[email] = password
    save_users(users)
    
    # Generate verification token
    token = uuid.uuid4().hex
    tokens = load_tokens()
    tokens[token] = f"verify:{email}" # Prefix to distinguish from reset password
    save_tokens(tokens)

    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            # Check if profile exists (shouldn't, but good practice)
            cursor.execute("SELECT * FROM user_profiles WHERE email = %s", (email,))
            if not cursor.fetchone():
                cursor.execute(
                    "INSERT INTO user_profiles (email, full_name, is_verified, total_credits, used_credits) VALUES (%s, %s, %s, %s, %s)",
                    (email, fullname, False, 1000, 0)
                )
                conn.commit()
            cursor.close()
            conn.close()
        
        send_verification_email(email, token)
        return "<h3 style='color:green;text-align:center;'>Registered! Please check your email to verify your account.</h3>"

    except Exception as e:
        print(f"Error creating user profile: {e}")
        return "<h3 style='color:red;text-align:center;'>Error creating account. Please try again.</h3>"


@app.route("/verify/<token>")
def verify_email(token):
    tokens = load_tokens()
    if token not in tokens or not tokens[token].startswith("verify:"):
        return "<h3 style='color:red;text-align:center;'>Invalid or expired verification link.</h3>"
    
    email = tokens.pop(token).split(":", 1)[1]
    save_tokens(tokens)
    
    # Update DB to set is_verified=True
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE user_profiles SET is_verified = 1 WHERE email = %s", (email,))
            conn.commit()
            cursor.close()
            conn.close()
    except Exception as e:
        print(f"Error verifying email: {e}")
        
    return send_file(os.path.join(os.path.dirname(__file__), 'verification-success.html'))


@app.route("/forgot-password")
def forgot_password_page():
    print("DEBUG: Hit forgot-password route")
    return send_file(os.path.join(os.path.dirname(__file__), 'forget-password.html'))


@app.route("/forgot-password", methods=["POST"])
def forgot_password():
    email = request.form.get("email")
    users = load_users()
    if email not in users:
        with open(os.path.join(os.path.dirname(__file__), 'forget-password.html'), 'r', encoding='utf-8') as tmpl_f:
            content = tmpl_f.read()
        return render_template_string(content, error="Email not found!")
    token = uuid.uuid4().hex
    tokens = load_tokens()
    tokens[token] = email
    save_tokens(tokens)
    try:
        send_reset_email(email, token)
        with open(os.path.join(os.path.dirname(__file__), 'forget-password.html'), 'r', encoding='utf-8') as tmpl_f:
            content = tmpl_f.read()
        return render_template_string(content, error="Reset link sent to your email!")
    except Exception:
        with open(os.path.join(os.path.dirname(__file__), 'forget-password.html'), 'r', encoding='utf-8') as tmpl_f:
            content = tmpl_f.read()
        return render_template_string(content, error="Email sending failed.")


@app.route("/reset-password/<token>")
def reset_password_page(token):
    tokens = load_tokens()
    if token not in tokens:
        return "<h3 style='color:red;text-align:center;'>Invalid or expired link.</h3>"
    return f"""
<html><body style="background:#0a0a0f;color:white;display:flex;justify-content:center;align-items:center;height:100vh;font-family:Arial;">
<div style="background:#4b0082;padding:40px;border-radius:20px;"><h2>Reset Password</h2>
<form method="post" action="/reset-password/{token}">
<input type="password" name="password" placeholder="New Password" required style="width:300px;padding:10px;">
<button type="submit" style="padding:10px;background:#9b5de5;color:white;border:none;border-radius:5px;">Update</button>
</form></div></body></html>
    """


@app.route("/reset-password/<token>", methods=["POST"])
def reset_password(token):
    tokens = load_tokens()
    if token not in tokens:
        return "<h3 style='color:red;text-align:center;'>Invalid token.</h3>"
    email = tokens.pop(token)
    save_tokens(tokens)
    users = load_users()
    users[email] = request.form.get("password")
    save_users(users)
    return send_file(os.path.join(os.path.dirname(__file__), 'passwordupdated.html'))


# ── Server-side OAuth state store (bypasses browser session cookie issues) ──
# Keys: random token stored in a short-lived cookie
# Values: {'state': ..., 'redirect_uri': ...}
_oauth_state_store = {}
_oauth_state_lock = threading.Lock()

def _cleanup_oauth_states():
    """Remove states older than 10 minutes."""
    cutoff = time.time() - 600
    with _oauth_state_lock:
        expired = [k for k, v in _oauth_state_store.items() if v.get('ts', 0) < cutoff]
        for k in expired:
            del _oauth_state_store[k]


@app.route('/login/google')
def google_login():
    import secrets
    from authlib.common.security import generate_token

    redirect_uri = url_for('google_authorize', _external=True)
    print(f"DEBUG: Generated Redirect URI: {redirect_uri}")

    # Generate state and a store_key; embed store_key INTO the state param
    # Google passes state back unchanged, so we can carry our key inside it
    # Format: "<store_key>:<authlib_state>"
    authlib_state = generate_token(32)
    store_key = secrets.token_hex(16)
    compound_state = f"{store_key}:{authlib_state}"

    _cleanup_oauth_states()
    with _oauth_state_lock:
        _oauth_state_store[store_key] = {
            'state': authlib_state,
            'ts': time.time()
        }

    # Build the Google authorization URL with our compound state
    resp = google.create_authorization_url(redirect_uri, state=compound_state, prompt='select_account')
    auth_url = resp['url']
    print(f"DEBUG: Auth URL built (store_key={store_key[:8]}...)")
    return redirect(auth_url)


@app.route('/login/google/authorized')
def google_authorize():
    try:
        redirect_uri = url_for('google_authorize', _external=True)
        incoming_compound_state = request.args.get('state', '')
        incoming_code = request.args.get('code', '')

        print(f"DEBUG: Callback — state={incoming_compound_state[:20]!r}, code={'YES' if incoming_code else 'NO'}")

        # Parse compound state: "<store_key>:<authlib_state>"
        if ':' not in incoming_compound_state:
            print(f"DEBUG: Invalid state format: {incoming_compound_state!r}")
            return redirect(url_for('home') + '?google_error=1')

        store_key, authlib_state = incoming_compound_state.split(':', 1)

        with _oauth_state_lock:
            stored = _oauth_state_store.pop(store_key, None)

        if not stored or stored.get('state') != authlib_state:
            print(f"DEBUG: State mismatch. stored={stored}, expected_state={authlib_state!r}")
            return redirect(url_for('home') + '?google_error=1')

        # Exchange authorization code for a token via direct POST to Google's token endpoint
        # (FlaskOAuth2App doesn't have fetch_token; we use requests directly since state is managed manually)
        import requests as _requests
        token_resp = _requests.post(
            'https://oauth2.googleapis.com/token',
            data={
                'code': incoming_code,
                'client_id': os.getenv('GOOGLE_CLIENT_ID'),
                'client_secret': os.getenv('GOOGLE_CLIENT_SECRET'),
                'redirect_uri': redirect_uri,
                'grant_type': 'authorization_code',
            }
        )
        token_data = token_resp.json()
        print(f"DEBUG: Token response status: {token_resp.status_code}, keys: {list(token_data.keys())}")
        if 'error' in token_data:
            print(f"DEBUG: Token error: {token_data}")
            return redirect(url_for('home') + '?google_error=1')

        access_token = token_data.get('access_token')

        # Fetch user info using the access token
        userinfo_resp = _requests.get(
            'https://www.googleapis.com/oauth2/v3/userinfo',
            headers={'Authorization': f'Bearer {access_token}'}
        )
        user_info = userinfo_resp.json()
        email = user_info['email']
        print(f"DEBUG: Google user info fetched: {email}")

        users = load_users()
        if email not in users or users[email] == "google_oauth_user":
            conn = get_db_connection()
            if conn:
                try:
                    cursor = conn.cursor()
                    cursor.execute("SELECT email FROM user_profiles WHERE email = %s", (email,))
                    if not cursor.fetchone():
                        cursor.execute(
                            "INSERT INTO user_profiles (email, full_name, profile_image, is_verified, total_credits, used_credits) VALUES (%s, %s, %s, %s, %s, %s)",
                            (email, user_info.get('name', ''), user_info.get('picture', ''), True, 1000, 0)
                        )
                        conn.commit()
                    else:
                        cursor.execute("UPDATE user_profiles SET is_verified = 1 WHERE email = %s", (email,))
                        conn.commit()
                    cursor.close()
                    conn.close()
                except Exception as e:
                    print(f"Error creating/updating profile for Google user: {e}")

            if email not in users:
                users[email] = "google_oauth_user"
                save_users(users)

        session['user'] = email
        session.modified = True
        print(f"DEBUG: Google login successful for {email}")
        return redirect(url_for('dashboard'))
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"Google login failed: {e}")
        print(tb)
        # Also write to file so we can read it
        try:
            with open('oauth_debug.log', 'w') as _lf:
                _lf.write(f"Error: {e}\n\n{tb}\n")
                _lf.write(f"\nRequest args: {dict(request.args)}\n")
                _lf.write(f"Store contents at error: {list(_oauth_state_store.keys())}\n")
        except Exception:
            pass
        return redirect(url_for('home') + '?google_error=1')





@app.route("/login", methods=["GET"])
def login_page():
    if "user" in session:
        return redirect(url_for("dashboard"))
    # Use render_template_string or send_file, but consistent with others
    with open(os.path.join(os.path.dirname(__file__), 'sign-in.html'), 'r', encoding='utf-8') as f:
        content = f.read()
    return render_template_string(content)


@app.route("/login", methods=["POST"])
def login():
    email = request.form.get("email", "").strip()
    password = request.form.get("password", "")
    users = load_users()
    print(f"DEBUG: Login attempt for user: '{email}'")
    
    # --- IMPROVED DEBUGGING ---
    with open(os.path.join(os.path.dirname(__file__), 'debug_log.txt'), 'a') as f:
        f.write(f"\n[{datetime.datetime.now()}] Login Attempt: {email}\n")
        
        if email not in users:
            f.write(f"Result: Email not found in users.json\n")
            with open(os.path.join(os.path.dirname(__file__), 'sign-in.html'), 'r', encoding='utf-8') as tmpl_f:
                content = tmpl_f.read()
            return render_template_string(content, error="Email not found. Please Sign Up.")
            
        if users.get(email) != password:
            f.write(f"Result: Password mismatch. Stored: {users.get(email)[:3]}*** Provided: {password[:3]}***\n")
            with open(os.path.join(os.path.dirname(__file__), 'sign-in.html'), 'r', encoding='utf-8') as tmpl_f:
                content = tmpl_f.read()
            return render_template_string(content, error="Incorrect password! Please try again.")

        # Check if email is verified
        is_verified = False
        conn = get_db_connection()
        
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT is_verified FROM user_profiles WHERE email = %s", (email,))
                res = cursor.fetchone()
                if res:
                    is_verified = res[0] # Boolean value
                cursor.close()
                conn.close()
            except Exception as e:
                f.write(f"DB Error checking verification: {e}\n")
                print(f"Error checking verification status: {e}")
        else:
            # Fallback: If DB is unreachable (local testing without DB), allow login
            # assuming users.json auth succeeded.
            print("DEBUG: DB unreachable. Bypassing verification check for login.")
            f.write("DEBUG: DB unreachable. Bypassing verification check.\n")
            is_verified = True
        
        if not is_verified:
             f.write(f"Result: Email found but NOT verified.\n")
             with open(os.path.join(os.path.dirname(__file__), 'sign-in.html'), 'r', encoding='utf-8') as tmpl_f:
                content = tmpl_f.read()
             return render_template_string(content, error="Email not verified. Please check your inbox.")

        f.write(f"Result: Success.\n")
        session["user"] = email
        if request.form.get("remember"):
            session.permanent = True
            print(f"DEBUG: Session set to permanent for {email}")
        print(f"DEBUG: Login successful for '{email}'")
        return redirect(url_for("dashboard"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

def get_user_display_info(email):
    # Default values
    display_name = email
    initial = email[0].upper() if email else "?"
    profile_image = None
    credits = {"total": 1000, "used": 0, "remaining": 1000}
    
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT full_name, profile_image, total_credits, used_credits FROM user_profiles WHERE email = %s", (email,))
            res = cursor.fetchone()
            if res:
                # Ensure we handle sqlite3.Row or MySQL dict result
                row = dict(res)
                if row.get('full_name'):
                    display_name = row['full_name']
                    # Use first letter of full name for initial
                    if display_name.strip():
                        initial = display_name.strip()[0].upper()
                if row.get('profile_image'):
                    profile_image = row['profile_image']
                
                # Credits
                t = row.get('total_credits', 1000) or 1000
                u = row.get('used_credits', 0) or 0
                credits = {"total": t, "used": u, "remaining": t - u}
            
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Error fetching user display info: {e}")
    
    return display_name, initial, profile_image, credits


@app.route("/dashboard")
def dashboard():
    user = session.get("user")
    print(f"DEBUG: Dashboard accessed at {datetime.datetime.now()}", flush=True)
    if not user:
        return redirect(url_for("home"))
    
    display_name, initial, profile_image, credits = get_user_display_info(user)
    
    stats = {
        'total_projects': 0,
        'total_checks': 0,
        'total_keywords': 0,
        'best_rank': 'N/A',
        'is_running': False
    }
    
    # Check for active task
    # Check for active task
    active_tasks = get_active_tasks_for_user(user)
    if active_tasks:
        stats['is_running'] = True
        stats['active_projects_count'] = len(active_tasks)
    else:
        stats['active_projects_count'] = 0
        
    recent_projects = []
    
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            
            # 1. Total Projects
            cursor.execute("SELECT COUNT(DISTINCT project_name) as cnt FROM runs WHERE user_email = %s AND project_name != 'No Project'", (user,))
            res = cursor.fetchone()
            if res: stats['total_projects'] = res['cnt']
            
            # 2. Total Websites (Unique)
            cursor.execute("SELECT COUNT(DISTINCT target_domain) as cnt FROM runs WHERE user_email = %s", (user,))
            res = cursor.fetchone()
            if res: stats['total_checks'] = res['cnt']
            
            # 3. Keywords monitored
            cursor.execute("SELECT COUNT(*) as cnt FROM results r JOIN runs ru ON r.run_id = ru.id WHERE ru.user_email = %s", (user,))
            res = cursor.fetchone()
            if res: stats['total_keywords'] = res['cnt']
            
            # 4. Best Rank (Numeric MIN, filtering out non-numeric/empty)
            cursor.execute("""
                SELECT MIN(CAST(rank AS SIGNED)) as best,
                       (SELECT keyword FROM results r2 JOIN runs ru2 ON r2.run_id = ru2.id 
                        WHERE ru2.user_email = %s AND r2.rank = (SELECT MIN(CAST(rank AS SIGNED)) FROM results r3 JOIN runs ru3 ON r3.run_id = ru3.id WHERE ru3.user_email = %s AND r3.rank REGEXP '^[0-9]+$')
                        LIMIT 1) as best_keyword
                FROM results r JOIN runs ru ON r.run_id = ru.id 
                WHERE ru.user_email = %s AND rank REGEXP '^[0-9]+$' AND rank != 'Not found in top 100'
            """, (user, user, user))
            res = cursor.fetchone()
            if res and res['best'] is not None: 
                stats['best_rank'] = f"#{res['best']}"
                stats['best_keyword'] = res['best_keyword']
            else:
                stats['best_keyword'] = None
            
            # 5. Recent Activity (Last 10)
            # Pagination Handling (AJAX aware)
            try:
                page = int(request.args.get('page', 1))
            except (ValueError, TypeError):
                page = 1
            if page < 1: page = 1
            per_page = 10
            offset = (page - 1) * per_page
            
            # Get total count first
            cursor.execute("SELECT COUNT(*) as count FROM runs WHERE user_email = %s", (user,))
            total_runs = cursor.fetchone()['count']
            import math
            total_pages = math.ceil(total_runs / per_page)
            if total_pages < 1: total_pages = 1

            cursor.execute("""
                SELECT r.id, r.timestamp, r.project_name, r.target_domain, r.excel_filename,
                        (SELECT COUNT(*) FROM results WHERE run_id = r.id) as keywords_count,
                        (SELECT COUNT(*) FROM results WHERE run_id = r.id AND rank REGEXP '^[0-9]+$' AND CAST(rank AS SIGNED) <= 10) as top_keywords
                FROM runs r
                WHERE r.user_email = %s
                ORDER BY r.timestamp DESC
                LIMIT %s OFFSET %s
            """, (user, per_page, offset))
            rows = cursor.fetchall()
            recent_projects = [dict(row) for row in rows]
            current_page = page
            for p in recent_projects:
                p['timestamp'] = parse_timestamp(p['timestamp'])
                # p['timestamp'] is now datetime obj or None or original string (if parse failed)

            # AJAX Request Check
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return JSON with rendered segments
                
                # Template for Rows
                rows_tmpl = """
                {% for run in recent_projects %}
                <tr style="border-bottom: 1px solid #eee;">
                    <td style="padding: 12px 10px; text-align: center;">{{ loop.index + (current_page - 1) * 10 }}</td>
                    <td style="padding: 12px 10px; text-align: left;" data-label="Project">{{ run.project_name }}</td>
                    <td style="padding: 12px 10px;" data-label="url"><a
                            href="{{ run.target_domain if 'http' in run.target_domain else 'https://' + run.target_domain }}"
                            target="_blank" title="{{ run.target_domain }}"
                            style="color:#00B29D; text-decoration:none;">{{ run.target_domain[:30] + '...' if run.target_domain|length > 30 else run.target_domain }}</a></td>
                    <td style="padding: 12px 10px; font-family: 'Lexend Deca', sans-serif; font-weight: 500; font-size: 13px; color: #666666;"
                        data-label="Date & Time">{{
                        run.timestamp.strftime('%d-%m-%Y, %I:%M %p') if
                        run.timestamp and
                        run.timestamp.strftime is defined else run.timestamp }}</td>
                    <td style="padding: 12px 10px;" data-label="Keywords Checked">{{ run.keywords_count }}</td>
                    <td style="padding: 12px 10px;" data-label="Keywords in Top 10">
                        <span style="font-weight: 600; color: #333;">{{ run.top_keywords }}</span>
                    </td>
                    <td data-label="Actions"
                        style="padding: 12px 10px; display: flex; gap: 15px; justify-content: center; align-items: center;">
                        <a href="javascript:void(0)" data-project-name="{{ run.project_name }}"
                            onclick="viewResults('{{ run.id }}', this.getAttribute('data-project-name'))"
                            title="View Quick Result">
                            <img src="./images/eye-img.svg" style="width: 20px;">
                        </a>
                        <a href="/download-result/{{ run.id }}" title="Download Excel">
                            <img src="./images/download-img.webp" style="width: 20px;">
                        </a>
                        <a href="javascript:void(0)" onclick="deleteRun('{{ run.id }}')"
                            title="Delete Run">
                            <img src="./images/delete-img.svg" style="width: 18px;">
                        </a>
                    </td>
                </tr>
                {% else %}
                <tr>
                    <td colspan="7">No recent activity found. Start a <a href="/rank-check">Rank Check</a>.</td>
                </tr>
                {% endfor %}
                """
                
                # Template for Pagination
                pag_tmpl = """
                {% if total_pages > 1 %}
                <div class="pagination"
                    style="display: flex; justify-content: center; gap: 10px; margin-top: 20px; align-items: center;">
                    {% if current_page > 1 %}
                    <a href="?page={{ current_page - 1 }}" onclick="loadDashboardPage(event, {{ current_page - 1 }})"
                        style="padding: 0 10px; height: 36px; display: inline-flex; justify-content: center; align-items: center; border: 1px solid #ddd; border-radius: 8px; color: #333; text-decoration: none; font-family: 'Lexend Deca', sans-serif; font-size: 14px; font-weight: 500; line-height: 20px;">&laquo;
                        Previous</a>
                    {% endif %}

                    {% for p in range(1, total_pages + 1) %}
                    {% if p == current_page %}
                    <span
                        style="padding: 8px 12px; border: 1px solid #00B29D; background-color: #00B29D; color: white; border-radius: 6px; font-size: 14px;">{{
                        p }}</span>
                    {% elif p == 1 or p == total_pages or (p >= current_page - 2 and p <= current_page + 2) or
                        (current_page <=4 and p <=5) or (current_page>= total_pages - 3 and p >= total_pages - 4) %}
                        <a href="?page={{ p }}" onclick="loadDashboardPage(event, {{ p }})"
                            style="padding: 8px 12px; border: 1px solid #ddd; border-radius: 6px; color: #333; text-decoration: none; font-size: 14px;">{{
                            p }}</a>
                        {% elif (p == 2 and current_page > 4) or (p == total_pages - 1 and current_page <
                            total_pages - 3) %} <span style="padding: 0 5px; color: #666;">...</span>
                            {% endif %}
                            {% endfor %}

                            {% if current_page < total_pages %} <a
                                href="?page={{ current_page + 1 }}" onclick="loadDashboardPage(event, {{ current_page + 1 }})"
                                style="padding: 0 10px; height: 36px; display: inline-flex; justify-content: center; align-items: center; border: 1px solid #ddd; border-radius: 8px; color: #333; text-decoration: none; font-family: 'Lexend Deca', sans-serif; font-size: 14px; font-weight: 500; line-height: 20px;">
                                Next &raquo;</a>
                                {% endif %}
                </div>
                {% endif %}
                """
                
                html_rows = render_template_string(rows_tmpl, recent_projects=recent_projects, current_page=current_page)
                html_pagination = render_template_string(pag_tmpl, total_pages=total_pages, current_page=current_page)
                
                return jsonify({'html_rows': html_rows, 'html_pagination': html_pagination})
                    
            # 6. Top Performing Websites Logic
            top_performing_sites = []
            try:
                # Get all runs for user to group by Target Domain (Website)
                cursor.execute("SELECT id, target_domain, timestamp FROM runs WHERE user_email = %s ORDER BY target_domain, timestamp DESC", (user,))
                all_runs = cursor.fetchall()
                
                project_runs_map = {}
                for r in all_runs:
                    # Use Target Domain as the key
                    domain_name = r['target_domain'] or "Unknown"
                    if domain_name not in project_runs_map:
                        project_runs_map[domain_name] = []
                    project_runs_map[domain_name].append(r['id'])
                
                # For each website, compare latest vs previous
                for domain_name, run_ids in project_runs_map.items():
                    if not run_ids: continue
                    
                    latest_run_id = run_ids[0]
                    # Check timestamp of latest run
                    # We need to find the run object from all_runs to get timestamp
                    latest_run_obj = next((r for r in all_runs if r['id'] == latest_run_id), None)
                    if latest_run_obj:
                        ts = parse_timestamp(latest_run_obj['timestamp'])
                        # If run is older than 30 days, skip it
                        if ts and (datetime.datetime.now() - ts).days > 30:
                            continue

                    prev_run_id = run_ids[1] if len(run_ids) > 1 else None
                    
                    # Get stats for latest run
                    cursor.execute("SELECT keyword, rank FROM results WHERE run_id = %s", (latest_run_id,))
                    latest_results = cursor.fetchall() # list of dicts
                    
                    if not latest_results: continue

                    # Map keywords to ranks
                    latest_map = {}
                    ranks_sum = 0
                    ranked_count = 0
                    
                    for row in latest_results:
                        k = row['keyword']
                        r = row['rank']
                        latest_map[k] = r
                        
                        # Avg Position Calc
                        if r and str(r).isdigit():
                            val = int(r)
                            ranks_sum += val
                            ranked_count += 1
                    
                    avg_position = int(round(ranks_sum / ranked_count)) if ranked_count > 0 else 0
                    
                    improved = 0
                    dropped = 0
                    
                    if prev_run_id:
                        cursor.execute("SELECT keyword, rank FROM results WHERE run_id = %s", (prev_run_id,))
                        prev_results = cursor.fetchall()
                        prev_map = {row['keyword']: row['rank'] for row in prev_results}
                        
                        for k, curr_rank in latest_map.items():
                            prev_rank = prev_map.get(k)
                            
                            def parse_r(v):
                                if v and str(v).isdigit(): return int(v)
                                return 101 # Treat Not Found as > 100
                            
                            c_val = parse_r(curr_rank)
                            p_val = parse_r(prev_rank)
                            
                            if c_val < p_val:
                                improved += 1
                            elif c_val > p_val:
                                dropped += 1
                    
                    # Format URL for display
                    formatted_name = domain_name
                    try:
                        if formatted_name.startswith("http://"): formatted_name = formatted_name[7:]
                        elif formatted_name.startswith("https://"): formatted_name = formatted_name[8:]
                        if formatted_name.startswith("www."): formatted_name = formatted_name[4:]
                        formatted_name = formatted_name.split('/')[0]
                        formatted_name = "www." + formatted_name
                    except:
                        pass

                    top_performing_sites.append({
                        'client_name': formatted_name,
                        'keywords': len(latest_results),
                        'improved': improved,
                        'dropped': dropped,
                        'avg_position': avg_position
                    })
                
                # Create a copy of the data for Low Performing calculation before we sort/slice 'top'
                all_sites_data = list(top_performing_sites)

                # Sort by Improved descending for Top Performing
                top_performing_sites.sort(key=lambda x: x['improved'], reverse=True)
                top_performing_sites = top_performing_sites[:10]

                # --- Low Performing Sites Logic ---
                # Sort by Dropped descending
                low_performing_sites = sorted(all_sites_data, key=lambda x: x['dropped'], reverse=True)
                low_performing_sites = low_performing_sites[:10]

            except Exception as e:
                print(f"Error calculating performing sites: {e}")

            cursor.close()
            conn.close()
        except Exception as err:
            print(f"Error fetching dashboard data: {err}")

    
    with open(os.path.join(os.path.dirname(__file__), 'dashboard.html'), 'r', encoding='utf-8') as f:
        html = f.read()
    
    return render_template_string(html, user_email=user, display_name=display_name, user_initial=initial, profile_image=profile_image, credits=credits, stats=stats, recent_projects=recent_projects, top_performing_sites=top_performing_sites, low_performing_sites=low_performing_sites, current_page=current_page, total_pages=total_pages)


@app.route("/projects")
def projects():
    user = session.get("user")
    if not user:
        return redirect(url_for("home"))
    
    display_name, initial, profile_image, credits = get_user_display_info(user)
    
    # 1. Get Active Tasks
    active_tasks_list = get_active_tasks_for_user(user)
    active_projects = {}
    
    # Filter out stale tasks (e.g., started but no update for > 30 mins, likely crashed)
    # Or just show them. The user wants "check if its terminated". 
    # For now, let's filter purely by status='started' which get_active_tasks_for_user does.
    # We will refine the display string.
    
    for t in active_tasks_list:
        p_name = t.get("project_name") or "No Project"
        if p_name not in active_projects:
            active_projects[p_name] = []
            
        total = len(t.get('keywords', []))
        processed = t.get('processed_count', 0)
        
        # Format: "Processing (5 / 10)"
        status_text = t.get('status')
        if status_text in ['started', 'started_real']:
             status_text = f"Processing ({processed})"
        elif status_text == 'waiting_for_socket':
             status_text = "Initializing..."
             
        t['display_status'] = status_text
        t['total_keywords'] = total
        
        active_projects[p_name].append(t)

    with open(os.path.join(os.path.dirname(__file__), 'projects.html'), 'r', encoding='utf-8') as f:
        html = f.read()
    
    return render_template_string(html, user_email=user, display_name=display_name, user_initial=initial, profile_image=profile_image, credits=credits, 
                                  active_projects=active_projects, completed_projects={})

# ... (rest of code)


@app.route("/re-check-edit/<int:run_id>")
@app.route("/re-check-modify/<int:run_id>")
def re_check_edit_page(run_id):
    user = session.get("user")
    if not user:
        return redirect(url_for("home"))
    
    display_name, initial, profile_image, credits = get_user_display_info(user)
    
    # Fetch run details to pre-fill
    conn = get_db_connection()
    run_data = {}
    keywords = ""
    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT id, project_name, target_domain FROM runs WHERE id = %s AND user_email = %s", (run_id, user))
            run_data = cursor.fetchone()
            
            if run_data:
                 # Fetch keywords
                cursor.execute("SELECT keyword FROM results WHERE run_id = %s", (run_id,))
                results = cursor.fetchall()
                # Ensure we handle list of dicts
                keyword_list = [row['keyword'] for row in results]
                keywords = "\n".join(keyword_list)
            
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Error fetching run for edit: {e}")

    if not run_data:
         return "Run not found or unauthorized", 404

    with open(os.path.join(os.path.dirname(__file__), 're-check-edit-v2.html'), 'r', encoding='utf-8') as f:
        html = f.read()
    
    return render_template_string(html, 
                                  user_email=user, 
                                  display_name=display_name, 
                                  user_initial=initial, 
                                  profile_image=profile_image, 
                                  credits=credits,
                                  run_id=run_id,
                                  project_name=run_data['project_name'],
                                  target_domain=run_data['target_domain'],
                                  keywords=keywords)


@app.route("/api/run-results/<int:run_id>")
def get_run_results(run_id):
    try:
        user = session.get("user")
        if not user:
            return jsonify({"error": "Unauthorized"}), 401

        conn = get_db_connection()
        if not conn:
            return jsonify({"error": "Database connection failed"}), 500

        cursor = conn.cursor(dictionary=True)
        
        # 1. Fetch Current Run Details
        cursor.execute("SELECT id, user_email, project_name, target_domain, timestamp FROM runs WHERE id = %s", (run_id,))
        current_run = cursor.fetchone()
        
        if not current_run:
             cursor.close(); conn.close()
             return jsonify({"error": "Run not found"}), 404
        
        # Verify ownership
        if str(current_run['user_email']) != str(user):
             cursor.close(); conn.close()
             return jsonify({"error": "Unauthorized access to run"}), 403

        # 1.5 Fetch ALL runs for this project (for dropdown)
        cursor.execute("""
            SELECT id, timestamp FROM runs 
            WHERE user_email = %s 
              AND project_name = %s 
              AND target_domain = %s 
            ORDER BY timestamp DESC
        """, (user, current_run['project_name'], current_run['target_domain']))
        all_project_runs = cursor.fetchall()

        # 2. Determine Previous Run for Comparison
        previous_run = None
        
        # Check if user requested a specific comparison
        compare_id = request.args.get('compare_id')
        
        if compare_id:
             # Look for the specific run requested
             for r in all_project_runs:
                 if str(r['id']) == str(compare_id):
                     previous_run = r
                     break
        else:
            # Default: Find the latest run that happened BEFORE this one
            for r in all_project_runs:
                if r['id'] < run_id: # assuming IDs are sequential/chronological is mostly safe, or comparing timestamps
                    previous_run = r
                    break
        
        # 1.6 Fetch Best Ranks for each keyword across ALL historical runs of this domain
        best_ranks_map = {}
        try:
            # Query MIN rank for all keywords under this domain for this user
            cursor.execute("""
                SELECT keyword, MIN(CAST(rank AS SIGNED)) as best_rank
                FROM results 
                WHERE run_id IN (
                    SELECT id FROM runs 
                    WHERE user_email = %s AND target_domain = %s
                )
                AND rank REGEXP '^[0-9]+$'
                GROUP BY keyword
            """, (user, current_run['target_domain']))
            best_rows = cursor.fetchall()
            for brow in best_rows:
                best_ranks_map[brow['keyword']] = brow['best_rank']
        except Exception as be:
            print(f"DEBUG: Error fetching best ranks: {be}")
            # Non-critical, continue without best_ranks if error
        
        # 3. Fetch Results for Current Run
        cursor.execute("SELECT keyword, `rank`, page, landing_page FROM results WHERE run_id = %s ORDER BY sort_order", (run_id,))
        current_results = cursor.fetchall()
        
        # 4. Fetch Results for Previous Run (if exists)
        previous_results_map = {}
        if previous_run:
            try:
                cursor.execute("SELECT keyword, `rank` FROM results WHERE run_id = %s", (previous_run['id'],))
                prev_rows = cursor.fetchall()
                for row in prev_rows:
                    previous_results_map[row['keyword']] = row['rank']
            except Exception as e:
                print(f"Error fetching previous results: {e}")

        cursor.close()
        conn.close()

        # 5. Process and Combine Data
        processed_data = []
        for row in current_results:
            keyword = row['keyword']
            current_rank = row['rank']
            # DB 'page' column is now strictly Page Number (or whatever was there for old runs)
            page_num = row['page'] 
            # DB 'landing_page' column is the URL (for new runs)
            # If 'landing_page' column doesn't exist in row (e.g. strict SQLite cursor without it?), handle it.
            # But we added it.
            landing_url_db = dict(row).get('landing_page')
            
            # Logic: If we have explicit landing_url, use it. 
            # If not (old runs), check if 'page' column happens to look like a URL (from my previous temporary fix if they ran it)
            # or just show "-"
            
            landing_display = "-"
            if landing_url_db:
                 landing_display = landing_url_db
            elif page_num and not str(page_num).isdigit() and ("http" in str(page_num) or "www" in str(page_num)):
                 # Catch transition data where URL was in page column
                 landing_display = page_num
            
            
            # Comparison Logic
            prev_rank_val = previous_results_map.get(keyword)
            
            # Helper to parse rank string to int
            def parse_rank(r):
                if not r: return None
                if str(r).isdigit(): return int(r)
                # Handle "1 (approx)" or similar if exists, but usually it's "1" or "Not found..."
                return None

            curr_int = parse_rank(current_rank)
            prev_int = parse_rank(prev_rank_val)
            
            change_val = 0
            change_label = "stable" # stable, improved, dropped, new
            
            if curr_int and prev_int:
                diff = prev_int - curr_int # Lower rank is better. Exp: Prev 10, Curr 5. Diff = 5 (Improved)
                change_val = diff
                if diff > 0: change_label = "improved"
                elif diff < 0: change_label = "dropped"
            elif curr_int and not prev_int:
                change_label = "new"
            elif not curr_int and prev_int:
                change_label = "dropped" # Dropped out of ranking completely
                
            processed_data.append({
                "keyword": keyword,
                "rank": current_rank,
                "previous_rank": prev_rank_val if prev_rank_val else "N/A",
                "change": change_val,
                "status": change_label,
                "best_rank": best_ranks_map.get(keyword),
                "landing_page": landing_display
            })

        # Helper for consistent date formatting
        def format_date_ddmmyyyy(ts):
            dt = parse_timestamp(ts)
            if dt:
                return dt.strftime("%d-%m-%Y")
            # Robust fallback for string dates like "YYYY-MM-DD HH:MM:SS..."
            if ts and isinstance(ts, str) and "-" in ts:
                try:
                    # Capture "YYYY-MM-DD" part
                    date_part = ts.split(' ')[0]
                    parts = date_part.split('-')
                    if len(parts) == 3:
                        if len(parts[0]) == 4: # YYYY-MM-DD
                             return f"{parts[2]}-{parts[1]}-{parts[0]}"
                        return date_part # Already in unknown format, return as is
                except:
                    pass
            return str(ts) if ts else None

        # Format available runs for dropdown
        available_runs = []
        for r in all_project_runs:
            date_str = format_date_ddmmyyyy(r['timestamp'])
            available_runs.append({
                "id": r['id'],
                "date": date_str
            })

        response = {
            "meta": {
                "project_name": current_run['project_name'],
                "domain": current_run['target_domain'],
                "date": format_date_ddmmyyyy(current_run['timestamp']),
                "previous_date": format_date_ddmmyyyy(previous_run['timestamp']) if previous_run else None,
                "previous_run_id": previous_run['id'] if previous_run else None,
                "available_runs": available_runs
            },
            "data": processed_data
        }
        
        return jsonify(response)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Server Error: {str(e)}"}), 500


@app.route("/history")
def history():
    user = session.get("user")
    if not user:
        return redirect(url_for("home"))
    
    display_name, initial, profile_image, credits = get_user_display_info(user)
    
    print(f"DEBUG: history route accessed by user: {user}")
    history_data = []
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            # Fetch all runs
            cursor.execute("""
                SELECT r.id, r.timestamp, r.project_name, r.target_domain, r.excel_filename, r.total_keywords,
                       (SELECT COUNT(*) FROM results WHERE run_id = r.id) as keywords_count,
                       (SELECT COUNT(*) FROM results WHERE run_id = r.id AND rank != 'Not Found' AND rank != 'N/A' AND rank != 'Not found in top 100') as top_keywords,
                       (SELECT MIN(CAST(rank AS SIGNED)) FROM results WHERE run_id = r.id AND rank REGEXP '^[0-9]+$' AND rank != 'Not found in top 100') as best_rank
                FROM runs r
                WHERE r.user_email = %s
                ORDER BY r.timestamp DESC
            """, (user,))
            rows = cursor.fetchall()
            # Group runs by project
            projects_map = {}
            for row in rows:
                row = dict(row)
                p_name = (row['project_name'] or "").strip() or "No Project"
                
                # Parse timestamp
                ts = parse_timestamp(row['timestamp'])
                row['timestamp_obj'] = ts # Keep obj for sorting
                
                # Keep as object for template formatting
                row['timestamp'] = ts 
                row['created_on'] = ts
                row['last_check'] = ts
                row['latest_run_id'] = row['id']
                
                if p_name not in projects_map:
                    projects_map[p_name] = []
                projects_map[p_name].append(row)
            
            # Sort within projects
            for p_name in projects_map:
                projects_map[p_name].sort(key=lambda x: x['timestamp_obj'] or datetime.datetime.min, reverse=True)
            
            # Sort projects
            sorted_projects = dict(sorted(
                projects_map.items(),
                key=lambda item: max((d['timestamp_obj'] or datetime.datetime.min) for d in item[1]),
                reverse=True
            ))
            
            history_data = sorted_projects
            
            print(f"DEBUG: fetched {len(rows)} runs, grouped into {len(history_data)} projects for user {user}")
            cursor.close()
            conn.close()
        except Exception as err:
            print(f"Error fetching history: {err}")
    else:
        print("DEBUG: Failed to connect to DB in history route")

    with open(os.path.join(os.path.dirname(__file__), 'history.html'), 'r', encoding='utf-8') as f:
        html = f.read()
    
    return render_template_string(html, user_email=user, display_name=display_name, user_initial=initial, profile_image=profile_image, credits=credits, completed_projects=history_data)


@app.route("/comparison", methods=['GET'])
def get_comparison_form():
    user = session.get("user")
    if not user:
        return redirect(url_for("home"))
    
    display_name, initial, profile_image, credits = get_user_display_info(user)
    
    projects = []
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT DISTINCT project_name FROM runs WHERE user_email = %s ORDER BY project_name", (user,))
            projects = [row['project_name'] for row in cursor.fetchall()]
            cursor.close()
        except Exception as err:
            print(f"Error fetching project list for comparison: {err}")
        finally:
            if conn.is_connected():
                conn.close()

    with open(os.path.join(os.path.dirname(__file__), 'comparison.html'), 'r', encoding='utf-8') as f:
        html = f.read()
    
    return render_template_string(html, user_email=user, display_name=display_name, user_initial=initial, profile_image=profile_image, credits=credits, projects=projects)

@app.route("/comparison", methods=['POST'])
def submit_comparison_form():
    user = session.get("user")
    if not user:
        return redirect(url_for("home"))

    project_name = request.form.get('project_name')
    keyword = request.form.get('keyword')
    target_domain = request.form.get('target_domain')
    from_date = request.form.get('from_date')
    to_date = request.form.get('to_date')

    ranking_trend = []
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            
            query = """
                SELECT ru.timestamp, res.rank, res.page
                FROM results res
                JOIN runs ru ON res.run_id = ru.id
                WHERE ru.user_email = %s AND res.keyword = %s
            """
            params = [user, keyword]

            if project_name:
                query += " AND ru.project_name = %s"
                params.append(project_name)
            
            if target_domain:
                # Remove common prefixes for better matching
                clean_domain = target_domain.replace('https://', '').replace('http://', '').replace('www.', '')
                query += " AND ru.target_domain LIKE %s"
                params.append(f"%{clean_domain}%")
            
            if from_date:
                query += " AND DATE(ru.timestamp) >= %s"
                params.append(from_date)

            if to_date:
                query += " AND DATE(ru.timestamp) <= %s"
                params.append(to_date)
            
            query += " ORDER BY ru.timestamp ASC"

            cursor.execute(query, tuple(params))
            ranking_trend = cursor.fetchall()
            cursor.close()
        except Exception as err:
            print(f"Error fetching ranking trend: {err}")
        finally:
            if conn.is_connected():
                conn.close()
    
    # Format timestamp for session serialization and display
    trend_dict = {}
    for row in ranking_trend:
        row_dict = dict(row)
        ts = parse_timestamp(row_dict.get('timestamp'))
        if isinstance(ts, datetime.datetime):
            date_key = ts.strftime('%Y-%m-%d')
            row_dict['timestamp'] = ts.isoformat()
            row_dict['display_date'] = ts.strftime('%d %b %Y')
        else:
            # Fallback for unexpected types
            date_key = str(row_dict.get('timestamp'))[:10]
            row_dict['timestamp'] = str(row_dict.get('timestamp'))
            row_dict['display_date'] = str(row_dict.get('timestamp'))
        
        # Keep only the latest entry for each day if there are multiple
        trend_dict[date_key] = row_dict
    
    # Generate full date range if dates are provided
    final_trend = []
    if from_date and to_date:
        try:
            start_dt = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
            end_dt = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
            
            curr = start_dt
            while curr <= end_dt:
                date_key = curr.strftime('%Y-%m-%d')
                if date_key in trend_dict:
                    final_trend.append(trend_dict[date_key])
                else:
                    final_trend.append({
                        'display_date': curr.strftime('%d %b %Y'),
                        'timestamp': curr.isoformat(),
                        'rank': None,
                        'page': None
                    })
                curr += timedelta(days=1)
        except Exception as e:
            print(f"Error generating date range: {e}")
            final_trend = list(trend_dict.values())
    else:
        # Fallback to just what we have if no dates specified
        final_trend = sorted(trend_dict.values(), key=lambda x: x['timestamp'])

    session['ranking_trend'] = final_trend
    session['comparison_params'] = {
        'project_name': project_name,
        'keyword': keyword,
        'target_domain': target_domain,
        'from_date': from_date,
        'to_date': to_date
    }
    return redirect(url_for("show_comparison_ranking"))


@app.route("/comparison-ranking", methods=['GET'])
def show_comparison_ranking():
    user = session.get("user")
    if not user:
        return redirect(url_for("home"))

    display_name, initial, profile_image, credits = get_user_display_info(user)
    ranking_trend = session.pop('ranking_trend', [])
    comparison_params = session.pop('comparison_params', {})

    projects = []
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT DISTINCT project_name FROM runs WHERE user_email = %s AND project_name != 'No Project' ORDER BY project_name", (user,))
            projects = [row['project_name'] for row in cursor.fetchall()]
            cursor.close()
        except Exception as err:
            print(f"Error fetching project list for comparison: {err}")
        finally:
            if conn.is_connected():
                conn.close()

    with open(os.path.join(os.path.dirname(__file__), 'comparison-ranking.html'), 'r', encoding='utf-8') as f:
        html = f.read()

    return render_template_string(html, user_email=user, display_name=display_name, user_initial=initial, profile_image=profile_image,
                                  projects=projects, ranking_trend=ranking_trend,
                                  comparison_params=comparison_params, credits=credits)


@app.route("/settings")
def settings():
    user = session.get("user")
    if not user:
        return redirect(url_for("home"))

    display_name, initial, profile_image, credits = get_user_display_info(user)
    
    profile = {'email': user, 'full_name': '', 'phone': '', 'profile_image': None}
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM user_profiles WHERE email = %s", (user,))
            res = cursor.fetchone()
            if res:
                profile = res
            else:
                profile['email'] = user
            cursor.close()
            conn.close()
        except Exception as err:
            print(f"Error fetching profile: {err}")

    with open(os.path.join(os.path.dirname(__file__), 'settings.html'), 'r', encoding='utf-8') as f:
        html = f.read()
    
    return render_template_string(html, user_email=user, display_name=display_name, user_initial=initial, profile_image=profile_image, credits=credits, profile=profile)


@app.route("/update-profile", methods=["POST"])
def update_profile():
    user = session.get("user")
    if not user:
        return redirect(url_for("home"))
        
    full_name = request.form.get("full_name", "").strip()
    phone = request.form.get("phone", "").strip()
    country_code = request.form.get("country_code", "").strip()
    
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            if conn.db_type == "mysql":
                cursor.execute("""
                    INSERT INTO user_profiles (email, full_name, phone, country_code)
                    VALUES (%s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE full_name=VALUES(full_name), phone=VALUES(phone), country_code=VALUES(country_code)
                """, (user, full_name, phone, country_code))
            else:
                cursor.execute("""
                    INSERT INTO user_profiles (email, full_name, phone, country_code)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(email) DO UPDATE SET full_name=excluded.full_name, phone=excluded.phone, country_code=excluded.country_code
                """, (user, full_name, phone, country_code))
            conn.commit()
            cursor.close()
            conn.close()
        except Exception as err:
            print(f"Error saving profile: {err}")
            
    return redirect(url_for("settings"))

@app.route('/change-password', methods=['POST'])
def change_password():
    user_email = session.get('user')
    if not user_email:
        return redirect(url_for('home'))

    # Strip just in case there are hidden spaces in session or form
    user_email = user_email.strip()
    current_password = request.form.get('current_password', "")
    new_password = request.form.get('new_password', "")
    confirm_password = request.form.get('confirm_password', "")

    users = load_users()
    stored_password = users.get(user_email)
    
    # Log to file for debugging
    debug_path = os.path.join(os.path.dirname(__file__), "debug_log.txt")
    with open(debug_path, "a") as f:
        f.write(f"\n--- Password Change Attempt ---\n")
        f.write(f"Session User: {repr(user_email)}\n")
        f.write(f"Users found: {list(users.keys())}\n")
        f.write(f"Stored Password: {repr(stored_password)}\n")
        f.write(f"Provided Password: {repr(current_password)}\n")
        f.write(f"Match: {stored_password == current_password}\n")

    if stored_password == current_password:
        users[user_email] = new_password
        save_users(users)
        
        add_notification(user_email, "You have successfully changed your password.")
        
        with open(os.path.join(os.path.dirname(__file__), 'passwordupdated.html'), 'r', encoding='utf-8') as f:
            html = f.read()
        return render_template_string(html)
    else:
        return f'<h3 style="color:red;text-align:center;">Incorrect current password. Your session user is {repr(user_email)}. <a href="/settings">Try again</a></h3>'

@app.route("/upload-profile-image", methods=["POST"])
def upload_profile_image():
    user = session.get("user")
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
        
    if file:
        # Create uploads directory if not exists
        upload_folder = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)
            
        # Generate unique filename to prevent caching/overwriting issues
        import uuid
        ext = os.path.splitext(file.filename)[1]
        filename = f"profile_{uuid.uuid4().hex}{ext}"
        filepath = os.path.join(upload_folder, filename)
        file.save(filepath)
        
        # Update database
        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor()
                # Upsert profile image
                if conn.db_type == "mysql":
                    cursor.execute("""
                        INSERT INTO user_profiles (email, profile_image)
                        VALUES (%s, %s)
                        ON DUPLICATE KEY UPDATE profile_image=VALUES(profile_image)
                    """, (user, filename))
                else:
                    cursor.execute("""
                        INSERT INTO user_profiles (email, profile_image)
                        VALUES (?, ?)
                        ON CONFLICT(email) DO UPDATE SET profile_image=excluded.profile_image
                    """, (user, filename))
                conn.commit()
                cursor.close()
                conn.close()
                return jsonify({"message": "Image uploaded successfully", "filename": filename})
            except Exception as err:
                print(f"Error saving profile image: {err}")
                return jsonify({"error": str(err)}), 500
    
    return jsonify({"error": "Upload failed"}), 500


@app.route('/delete-profile-image', methods=['POST'])
def delete_profile_image():
    user = session.get('user')
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            # Set profile_image to NULL or empty
            cursor.execute("UPDATE user_profiles SET profile_image = NULL WHERE email = %s", (user,))
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"success": True})
    except Exception as e:
        print(f"Error deleting profile image: {e}")
        return jsonify({"error": str(e)}), 500
            
    return jsonify({"error": "DB Connection failed"}), 500


@app.route('/delete-account', methods=['POST'])
def delete_account():
    user_email = session.get('user')
    if not user_email:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401

    try:
        # Delete from database
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            # The ON DELETE CASCADE in the results table will handle associated results
            cursor.execute("DELETE FROM runs WHERE user_email = %s", (user_email,))
            cursor.execute("DELETE FROM user_profiles WHERE email = %s", (user_email,))
            cursor.execute("DELETE FROM notifications WHERE user_email = %s", (user_email,))
            conn.commit()
            cursor.close()
            conn.close()

        # Delete from users.json
        users = load_users()
        if user_email in users:
            del users[user_email]
            save_users(users)

        # Clear session
        session.pop('user', None)

        return jsonify({'success': True})
    except Exception as e:
        print(f"Error deleting account for {user_email}: {e}")
        return jsonify({'success': False, 'error': 'An internal error occurred.'}), 500


@app.route("/rank-check")
def rank_check_page():
    user = session.get("user")
    if not user:
        return redirect(url_for("home"))
    
    display_name, initial, profile_image, credits = get_user_display_info(user)
    
    # Fetch previous project names for suggestions
    previous_projects = []
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT project_name FROM runs WHERE user_email = %s AND project_name IS NOT NULL AND project_name != '' AND project_name != 'nan' ORDER BY timestamp DESC LIMIT 20", (user,))
            rows = cursor.fetchall()
            previous_projects = [row[0] for row in rows]
            cursor.close()
            conn.close()
    except Exception as e:
        print(f"Error fetching project names: {e}")

    with open(os.path.join(os.path.dirname(__file__), 'rank-check.html'), 'r', encoding='utf-8') as f:
        html = f.read()
    
    # Remove active task check for UI blocking
    # current_task = get_task_for_user(user) ... 
    
    return render_template_string(html, user_email=user, display_name=display_name, user_initial=initial, profile_image=profile_image, credits=credits, previous_projects=previous_projects)


@app.route("/start-task", methods=["POST"])
def start_task():
    user = session.get("user")
    if not user:
        return redirect(url_for("home"))

    # Only one task at a time check
    # Removed single task check
    # current_task = get_task_for_user(user) ...

    keywords = []
    if "file" in request.files and request.files["file"].filename:
        file = request.files["file"]
        # Save file to disk first
        upload_folder = os.path.join(STATIC_FOLDER, 'uploads')
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)
        
        filename = f"keywords_{int(time.time())}_{file.filename}"
        filepath = os.path.join(upload_folder, filename)
        file.save(filepath)
        print(f"DEBUG: Saved uploaded keywords file to {filepath}")
        
        # Read from the saved file
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        keywords = [line.strip() for line in content.splitlines() if line.strip()]

    manual = request.form.get("keywords", "")
    if manual:
        # Support both comma and newline separation in backend just in case
        manual_keywords = []
        for part in manual.split(","):
            for line in part.splitlines():
                if line.strip():
                    manual_keywords.append(line.strip())
        keywords.extend(manual_keywords)

    if not keywords:
        return "<h3 style='color:red;text-align:center;'>No keywords provided!</h3>"

    target_domain = request.form.get("target_domain", "").strip()
    if not target_domain:
        return "<h3 style='color:red;text-align:center;'>Target domain required!</h3>"

    max_pages = max(1, min(50, int(request.form.get("max_pages", 10))))
    location = request.form.get("location", "").strip()
    
    # STRICT LOCATION REQUIREMENT
    if not location or location.lower() in ["global", ""]:
        return "<h3 style='color:red;text-align:center;'>Validation Error: Location is MANDATORY. Please select a specific country to ensure accurate proxy targeting.</h3>"
        
    project_name = request.form.get("project_name", "No Project").strip()
    if not project_name or project_name.lower() == 'nan':
        project_name = "No Project"
    run_id = None
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            
            # Check Credits Logic
            required_credits = len(keywords)
            cursor.execute("SELECT total_credits, used_credits FROM user_profiles WHERE email = %s", (user,))
            row = cursor.fetchone()
            if not row:
                cursor.close(); conn.close()
                return "<h3 style='color:red;text-align:center;'>User profile not found</h3>"
            
            total = row[0] 
            used = row[1]
            if isinstance(row, dict):
                 total = row.get('total_credits', 0)
                 used = row.get('used_credits', 0)
            
            if (total - used) < required_credits:
                 cursor.close(); conn.close()
                 return f"<h3 style='color:red;text-align:center;'>Insufficient credits. You have {total-used} remaining but need {required_credits}.</h3>"
            
            # Deduct Credits
            cursor.execute("UPDATE user_profiles SET used_credits = used_credits + %s WHERE email = %s", (required_credits, user))

            cursor.execute("""
                INSERT INTO runs (user_email, project_name, target_domain, location, max_pages, total_keywords, excel_filename, excel_data, timestamp)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (user, project_name, target_domain, location, max_pages, len(keywords), None, None, datetime.datetime.now()))
            conn.commit()
            run_id = cursor.lastrowid
            cursor.close()
            conn.close()
    except Exception as e:
        print(f"Error creating initial run record: {e}")
        return f"<h3 style='color:red;'>Database Error: {e}</h3>"

    print(f"DEBUG: start_task called for user {user}. Keywords: {len(keywords)}, Location: {location}")
    if not run_id:
         print("DEBUG: Failed to generate run ID inside start_task")
         return "<h3 style='color:red;'>Failed to create run ID. Check server logs.</h3>"

    task_obj = {
        "user": user,
        "keywords": keywords,
        "target_domain": target_domain,
        "location": location,
        "project_name": project_name,
        "max_pages": max_pages,
        "sid": None,
        "status": "waiting_for_socket",
        "cancelled": False,
        "run_id": run_id       # Store DB ID
    }

    # Save task keyed by run_id, not user
    save_task(run_id, task_obj)

    return redirect(url_for("progress", run_id=run_id))


@app.route("/progress/<int:run_id>")
def progress(run_id):
    user = session.get("user")
    if not user:
        return redirect(url_for("home"))

    task = get_task(run_id)
    if not task:
        # Check if it was completed and removed from tasks.json, or just invalid
        # For now, just show error
        return render_template_string(f"<h3 style='color:red;text-align:center;'>Task not found or expired. <a href='/dashboard'>Go back</a></h3>")
    
    if task.get('user') != user:
        return "<h3 style='color:red;'>Unauthorized</h3>"

    display_name, initial, profile_image, credits = get_user_display_info(user)
    target_domain = task["target_domain"]

    # Fetch previous project names for suggestions (needed for rank-check.html background)
    previous_projects = []
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT project_name FROM runs WHERE user_email = %s AND project_name IS NOT NULL AND project_name != '' AND project_name != 'nan' ORDER BY timestamp DESC LIMIT 20", (user,))
            rows = cursor.fetchall()
            previous_projects = [row[0] for row in rows]
            cursor.close()
            conn.close()
    except Exception as e:
        print(f"Error fetching project names: {e}")

    # Load rank-check.html to use as the background
    try:
        base_html_path = os.path.join(os.path.dirname(__file__), 'rank-check.html')
        with open(base_html_path, 'r', encoding='utf-8') as f:
            base_html = f.read()
    except Exception as e:
        print(f"Error reading rank-check.html: {e}")
        return f"Error loading background template: {e}"

    # Define the Modal Overlay HTML
    modal_html = f"""
    <div id="progressOverlay" class="progress-overlay-container">
        <div class="progress-modal-box" style="position: relative;">
            <a href="/rank-check" class="progress-close-btn" title="Close">&times;</a>
            <h3 id="loadingTitle">Processing Your Keyword Check</h3>
            <p class="subtext" id="loadingSubtitle">
                We’re analyzing <strong style="color: #0f172a;" title="{target_domain}">{target_domain[:30] + '...' if len(target_domain) > 30 else target_domain}</strong>
            </p>

            <div class="progress-bar-wrapper">
                <div class="progress-bar-fill" id="progressBar">0%</div>
            </div>

            <p class="progress-status-text" id="loadingStep">Initializing...</p>
            
            <div class="progress-modal-buttons">
                <div class="progress-btn-row" id="controlBtns" style="justify-content: center;">
                    <button id="stopBtn" onclick="stopProgressTask()" class="action-btn-link btn-danger-solid" style="width: 156px !important; height: 36px !important; display: inline-flex; justify-content: center; align-items: center; padding: 0 !important; font-family: 'Lexend Deca', sans-serif; font-size: 13px; font-weight: 500; line-height: 20px; border-radius: 8px; flex: none;">Stop Project</button>
                </div>
                
                <a href="/download-result/{run_id}" class="action-btn-link btn-teal-solid" id="downloadBtn" style="display: none; width: 156px !important; height: 36px !important; justify-content: center; align-items: center; padding: 0 !important; font-family: 'Lexend Deca', sans-serif; font-size: 13px; font-weight: 500; line-height: 20px; border-radius: 8px; margin: 0 auto; flex: none;">Download Results</a>
            </div>
        </div>
    </div>
    """

    # Define Modal CSS
    modal_css = """
    <style>
        .progress-overlay-container {
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0, 0, 0, 0.5); /* Standard dark overlay */
            backdrop-filter: none;
            -webkit-backdrop-filter: none;
            z-index: 10000;
            display: flex;
            justify-content: center;
            align-items: center;
            padding: 24px;
        }

        .progress-modal-box {
            background: #ffffff;
            padding: 40px;
            border-radius: 24px;
            box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.25);
            text-align: center;
            max-width: 520px;
            width: 100%;
            border: 1px solid #e2e8f0;
            animation: modalFadeIn 0.3s ease-out;
        }

        @keyframes modalFadeIn {
            from { opacity: 0; transform: translateY(20px); }
            to { opacity: 1; transform: translateY(0); }
        }

        .progress-close-btn {
            position: absolute;
            top: 20px;
            right: 20px;
            font-size: 28px;
            color: #94a3b8;
            text-decoration: none;
            line-height: 1;
            transition: color 0.2s;
            cursor: pointer;
        }

        .progress-close-btn:hover {
            color: #475569;
        }

        .progress-modal-box h3 {
             color: #0f172a;
             font-size: 24px;
             margin-bottom: 12px;
             font-weight: 700;
        }

        .progress-modal-box p.subtext {
             color: #64748b;
             font-size: 15px;
             line-height: 1.6;
             margin-bottom: 32px;
        }

        .progress-bar-wrapper {
            background: #f1f5f9;
            border-radius: 99px;
            padding: 4px;
            margin-bottom: 24px;
            height: 36px;
            overflow: hidden;
            box-shadow: inset 0 2px 4px rgba(0,0,0,0.05);
        }
        
        .progress-bar-fill {
            height: 100%;
            border-radius: 99px;
            background: linear-gradient(90deg, #00B29D 0%, #F2BD2C 100%);
            width: 0%;
            transition: width 0.4s cubic-bezier(0.4, 0, 0.2, 1);
            display: flex; 
            align-items: center; 
            justify-content: center;
            color: #fff;
            font-weight: 700;
            font-size: 14px;
            text-shadow: 0 1px 2px rgba(0,0,0,0.1);
        }
        
        .progress-status-text {
            color: #475569;
            font-size: 15px;
            font-weight: 500;
            margin-bottom: 32px;
            min-height: 24px;
        }

        .progress-modal-buttons {
            display: flex;
            flex-direction: column;
            gap: 16px;
            width: 100%;
        }

        .progress-btn-row {
            display: flex;
            gap: 16px;
            width: 100%;
        }

        .action-btn-link {
            padding: 14px 24px;
            border-radius: 12px;
            font-weight: 600;
            cursor: pointer;
            text-decoration: none;
            font-size: 15px;
            text-align: center;
            border: none;
            transition: all 0.2s;
            flex: 1;
            font-family: 'Lexend Deca', sans-serif;
            display: inline-block;
        }

        .btn-gray-muted { background: #f1f5f9; color: #475569; }
        .btn-gray-muted:hover { background: #e2e8f0; }
        
        .btn-danger-solid { background: #ef4444; color: white; }
        .btn-danger-solid:hover { background: #dc2626; }
        
        .btn-teal-outline {
            background: transparent;
            color: #00B29D;
            border: 1.5px solid #00B29D;
        }
        .btn-teal-outline:hover { background: #f0fdfb; }
        
        .btn-teal-solid {
            background: #00B29D;
            color: white;
        }
        .btn-teal-solid:hover { background: #0d9488; }
    </style>
    """

    # Socket JS
    socket_js = f"""
    <script src="https://cdnjs.cloudflare.com/ajax/libs/socket.io/4.0.1/socket.io.js"></script>
    <script>
        (function() {{
            const socket = io();
            const runId = {run_id};
            let isStopped = false;

            socket.on('connect', () => {{
                 console.log("Progress socket connected");
                 socket.emit('join_task', {{ user: "{user}", run_id: runId }});
            }});
            
            socket.on('task_ready', data => {{
                if (String(data.run_id) === String(runId)) {{
                    socket.emit('start_rank_check', {{ run_id: runId }});
                }}
            }});

            socket.on('status_update', (data) => {{
                if (isStopped) return;
                
                if (data.msg) document.getElementById('loadingStep').innerText = data.msg;
                
                if (data.progress !== undefined) {{
                    const fill = document.getElementById('progressBar');
                    if (fill) {{
                        fill.style.width = data.progress + '%';
                        fill.innerText = data.progress + '%';
                    }}
                    
                    if (data.progress >= 100) {{
                        completeProgressTask();
                    }}
                }}
                
                if (data.error) {{
                    document.getElementById('loadingStep').innerText = "Error: " + data.error;
                    document.getElementById('loadingStep').style.color = "#ef4444";
                    if (data.error.includes("stopped") || data.error.includes("cancelled")) {{
                        isStopped = true;
                    }}
                }}
                
                if (data.done) {{
                    completeProgressTask();
                }}
            }});

            window.completeProgressTask = function() {{
                const controls = document.getElementById('controlBtns');
                if(controls) controls.style.display = 'none';
                
                const dlBtn = document.getElementById('downloadBtn');
                if(dlBtn) dlBtn.style.display = 'inline-flex';
                
                document.getElementById('loadingTitle').innerText = "Task Completed!";
                document.getElementById('loadingSubtitle').innerText = "Your keyword ranking report is ready.";
                document.getElementById('loadingStep').innerText = "Success! You can now download the report.";
                document.getElementById('loadingStep').style.color = "#10b981";
                
                const fill = document.getElementById('progressBar');
                if(fill) fill.style.background = "#10b981";
            }};

            window.stopProgressTask = function() {{
                if (typeof showConfirm === 'function') {{
                    showConfirm("Stop Task?", "Are you sure you want to stop this task?", () => {{
                        executeStop();
                    }});
                }} else {{
                    if (confirm("Are you sure you want to stop this task?")) {{
                        executeStop();
                    }}
                }}

                function executeStop() {{
                    isStopped = true;
                    const btn = document.getElementById('stopBtn');
                    if(btn) btn.innerText = "Stopping...";
                    
                    fetch("/api/stop-task/" + runId, {{ method: "POST" }})
                    .then(res => res.json())
                    .then(data => {{
                        document.getElementById('loadingStep').innerText = "Task stopped. Returning to form...";
                        document.getElementById('loadingStep').style.color = "#ef4444";
                        setTimeout(() => {{ window.location.href = "/rank-check"; }}, 1000);
                    }})
                    .catch(err => {{
                        console.error("Stop error:", err);
                        isStopped = false;
                        if(btn) btn.innerText = "Stop Project";
                    }});
                }}
            }};
        }})();
    </script>
    """

    # Inject into base HTML
    injected_html = base_html.replace('<head>', '<head><base href="/">')
    injected_html = injected_html.replace('</head>', modal_css + '</head>')
    injected_html = injected_html.replace('</body>', modal_html + socket_js + '</body>')

    return render_template_string(injected_html, 
                                  user_email=user, 
                                  display_name=display_name, 
                                  user_initial=initial, 
                                  profile_image=profile_image, 
                                  credits=credits, 
                                  previous_projects=previous_projects,
                                  task_running=False)


# ====================== SOCKET.IO EVENTS ======================

@socketio.on("join_task")
def handle_join_task(data):
    print(f"DEBUG: handle_join_task received for run_id {data.get('run_id')} from user {data.get('user')}")
    user_in_payload = data.get("user")
    user_in_session = session.get("user")
    run_id = data.get("run_id")

    # Only allow if session user matches payload user
    if not user_in_session or user_in_payload != user_in_session:
        emit("status_update", {"error": "Unauthorized"}, room=request.sid)
        return

    task = get_task(run_id)
    if not task:
        # Check if completed in DB but cleaned from RAM
        # For now, just error
        emit("status_update", {"error": "Task not found (expired?)"}, room=request.sid)
        return

    if task.get("cancelled"):
        emit("status_update", {"error": "Task was cancelled."}, room=request.sid)
        return

    # Allow attaching if not started yet, already started, OR completed
    if task.get("status") not in ["waiting_for_socket", "ready_to_start", "started", "completed"]:
        emit(
            "status_update",
            {"error": "Task finished or failed."},
            room=request.sid,
        )
        return

    # If task is already completed, show success immediately
    if task.get("status") == "completed":
        run_id = task.get("run_id")
        filename = task.get("filename", f"results_{run_id}.xlsx")
        emit(
            "status_update",
            {
                "done": True,
                "msg": "Completed! Download ready.",
                "filename": filename,
                "run_id": run_id,
                "progress": 100,
            },
            room=request.sid,
        )
        return

    # Attach socket to task and mark ready (or keep started)
    # If it was "started", we keep it "started" to avoid resetting logic, but we update SID?
    # Actually, we rely on room now, so SID is less critical, but good for debug.
    task["sid"] = request.sid
    if task["status"] != "started":
         task["status"] = "ready_to_start"
    
    save_task(run_id, task)
    
    # Join specific task room
    join_room(f"task_{run_id}")

    emit(
        "status_update",
        {"msg": "Connected! " + ("Resuming task..." if task["status"] == "started" else "Preparing to start rank check...")},
        room=request.sid,
    )
    # Notify client that it can now send start_rank_check (only if not already running)
    if task["status"] != "started":
         emit("task_ready", {"ok": True, "run_id": run_id}, room=request.sid)

@app.route("/api/stop-task/<int:run_id>", methods=['POST'])
def api_stop_task(run_id):
    user = session.get("user")
    if not user:
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    task = get_task(run_id)
    if not task:
        return jsonify({"success": True, "message": "Task already stopped or not found"})
         
    if task.get("user") != user:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    # Delete the task from the store immediately
    try:
        delete_task_from_disk(run_id)
    except Exception as e:
        print(f"Error deleting stopped task: {e}")

    # Also emit via socket if connected
    socketio.emit(
        "status_update",
        {"error": "Task stopped by user"},
        room=f"task_{run_id}"
    )
            
    return jsonify({"success": True})

@app.route("/api/cancel-task", methods=['POST'])
def api_cancel_task():
    user = session.get("user")
    if not user:
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    data = request.get_json() or {}
    run_id = data.get('run_id')
    
    if not run_id:
         return jsonify({"success": False, "error": "Missing run_id"}), 400

    task = get_task(run_id)
    if not task:
         return jsonify({"success": False, "error": "Task not found"}), 404
         
    if task.get("user") != user:
         return jsonify({"success": False, "error": "Unauthorized"}), 403

    # Delete the task from the store immediately
    try:
        delete_task_from_disk(run_id)
    except Exception as e:
        print(f"Error deleting cancelled task: {e}")

    # Also emit via socket if connected
    socketio.emit(
        "status_update",
        {"error": "Task cancelled by user"},
        room=f"task_{run_id}"
    )
            
    return jsonify({"success": True})

def start_task_internal(keywords, target_domain, location, project_name, max_pages, user_email, explicit_cost=None):
    """
    Helper to start a task programmatically.
    Returns new_run_id on success, raises Exception on failure.
    """
    import datetime
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            # Check Credits Logic
            if explicit_cost is not None:
                required_credits = int(explicit_cost)
            else:
                required_credits = len(keywords)

            cursor.execute("SELECT total_credits, used_credits FROM user_profiles WHERE email = %s", (user_email,))
            row = cursor.fetchone()
            if not row:
                cursor.close(); conn.close()
                raise Exception("User profile not found")
            
            total = row[0] # SQLite Row by index or column name
            used = row[1]
            if isinstance(row, dict):
                 total = row.get('total_credits', 0)
                 used = row.get('used_credits', 0)
            
            if (total - used) < required_credits:
                 cursor.close(); conn.close()
                 raise Exception(f"Insufficient credits. You have {total-used} remaining but need {required_credits}.")
            
            # Deduct Credits
            cursor.execute("UPDATE user_profiles SET used_credits = used_credits + %s WHERE email = %s", (required_credits, user_email))

            sql = """
                INSERT INTO runs (user_email, project_name, target_domain, location, max_pages, total_keywords, excel_filename, excel_data, timestamp)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(sql, (user_email, project_name, target_domain, location, max_pages, len(keywords), None, None, datetime.datetime.now()))
            conn.commit()
            run_id = cursor.lastrowid
            cursor.close()
            conn.close()
            
            if not run_id:
                raise Exception("Failed to generate run ID")

            task_obj = {
                "user": user_email,
                "keywords": keywords,
                "target_domain": target_domain,
                "location": location,
                "project_name": project_name,
                "max_pages": max_pages,
                "sid": None,
                "status": "waiting_for_socket",
                "cancelled": False,
                "run_id": run_id
            }

            save_task(run_id, task_obj)
            return run_id
    except Exception as e:
        print(f"Error creating run record in start_task_internal: {e}")
        raise e
    raise Exception("DB Connection failed")

@app.route("/api/run-results/<int:run_id>")
def api_run_results(run_id):
    user = session.get("user")
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
        
    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "DB Connection failed"}), 500

    try:
        cursor = conn.cursor(dictionary=True)
        # Verify ownership & Fetch Meta
        cursor.execute("SELECT user_email, project_name, target_domain, timestamp, location FROM runs WHERE id = %s", (run_id,))
        run = cursor.fetchone()
        
        if not run:
             cursor.close(); conn.close()
             return jsonify({"error": "Run not found"}), 404
             
        if run['user_email'] != user:
             cursor.close(); conn.close()
             return jsonify({"error": "Unauthorized"}), 403

        # Fetch Previous Runs for Comparison
        cursor.execute("SELECT id, timestamp FROM runs WHERE user_email = %s AND target_domain = %s AND project_name = %s ORDER BY timestamp DESC", 
                       (user, run['target_domain'], run['project_name']))
        available_runs_raw = cursor.fetchall()
        
        # Determine previous run (simple logic: specific comparison or immediate previous)
        compare_id = request.args.get('compare_id')
        previous_run_id = None
        previous_date = None
        
        available_runs = []
        for r in available_runs_raw:
            ts = parse_timestamp(r['timestamp'])
            date_str = ts.strftime('%Y-%m-%d %H:%M') if ts else str(r['timestamp'])
            available_runs.append({
                'id': r['id'],
                'date': date_str,
                'timestamp': ts.isoformat() if ts else str(r['timestamp'])
            })
            
            # Logic to find comparison run
            if compare_id and str(r['id']) == str(compare_id):
                previous_run_id = r['id']
                previous_date = date_str
        
        # If no specific compare_id, default to the one before current (if current is in list)
        if not previous_run_id and len(available_runs) > 1:
            # Find current run index
            for i, r in enumerate(available_runs):
                if str(r['id']) == str(run_id) and i + 1 < len(available_runs):
                    previous_run_id = available_runs[i+1]['id']
                    previous_date = available_runs[i+1]['date']
                    break

        # Fetch Current Results
        cursor.execute("SELECT keyword, page, rank, landing_page FROM results WHERE run_id = %s ORDER BY sort_order", (run_id,))
        current_results = cursor.fetchall()

        # Fetch Previous Results (if any)
        previous_results_map = {}
        if previous_run_id:
            cursor.execute("SELECT keyword, rank FROM results WHERE run_id = %s", (previous_run_id,))
            prev_rows = cursor.fetchall()
            for row in prev_rows:
                previous_results_map[row['keyword']] = row['rank']

        cursor.close()
        conn.close()

        # Merge Data
        merged_data = []
        for row in current_results:
            curr_rank = row['rank']
            prev_rank = previous_results_map.get(row['keyword'], "N/A")
            
            # Calculate change
            change = 0
            status = 'stable'
            
            # Helper to parse rank
            def parse_rank(r):
                if r is None or r == "N/A" or r == "Not found in top 100": return 101
                try: return int(r)
                except: return 101

            c_val = parse_rank(curr_rank)
            p_val = parse_rank(prev_rank)
            
            if prev_rank == "N/A":
                status = 'new'
            elif c_val < p_val:
                status = 'improved'
                change = p_val - c_val
            elif c_val > p_val:
                status = 'dropped'
                change = c_val - p_val
            else:
                status = 'stable'

            merged_data.append({
                'keyword': row['keyword'],
                'rank': curr_rank,
                'previous_rank': prev_rank,
                'change': change,
                'status': status,
                'landing_page': row['landing_page'],
                'page': row['page']  # Include Page Number
            })

        # Construct Meta
        ts = parse_timestamp(run['timestamp'])
        meta = {
            'project_name': run['project_name'],
            'domain': run['target_domain'],
            'date': ts.strftime('%Y-%m-%d %H:%M') if ts else str(run['timestamp']),
            'previous_run_id': previous_run_id,
            'previous_date': previous_date,
            'available_runs': available_runs
        }

        return jsonify({'meta': meta, 'data': merged_data})
    except Exception as e:
        if conn: conn.close()
        print(f"Error fetching results: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/re-check/<int:run_id>", methods=["POST"])
def api_re_check(run_id):
    user = session.get("user")
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM runs WHERE id = %s", (run_id,))
    run = cursor.fetchone()
    
    if not run:
        cursor.close(); conn.close()
        return jsonify({"error": "Original run not found"}), 404
        
    # Check ownership
    if run["user_email"] != user:
        cursor.close(); conn.close()
        return jsonify({"error": "Unauthorized"}), 403

    # Get keywords from Request (if provided) or Results (fallback)
    custom_keywords = request.json.get("keywords") if request.json else None

    if custom_keywords:
        if not isinstance(custom_keywords, list):
             return jsonify({"error": "Keywords must be a list"}), 400
        keywords = custom_keywords
    else:
        cursor.execute("SELECT keyword FROM results WHERE run_id = %s ORDER BY sort_order", (run_id,))
        results = cursor.fetchall()
        if not results:
             cursor.close(); conn.close()
             return jsonify({"error": "No keywords found in history to re-check."}), 400
        keywords = [r["keyword"] for r in results]
    
    cursor.close()
    conn.close()

    if not keywords:
        return jsonify({"error": "Keyword list is empty."}), 400

    # Remove duplicates and empty strings
    keywords = list(dict.fromkeys([k for k in keywords if k and str(k).strip()]))

    project_name = run["project_name"]
    target_domain = run["target_domain"]
    location = run["location"]
    max_pages = run["max_pages"] or 10
    
    # Start new task
    try:
        new_run_id = start_task_internal(
            keywords=keywords,
            target_domain=target_domain,
            location=location,
            project_name=project_name,
            max_pages=max_pages,
            user_email=user
        )
        return jsonify({"success": True, "run_id": new_run_id})
    except Exception as e:
        print(f"Error restarting task: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/remove-task", methods=['POST'])
def api_remove_task():
    user = session.get("user")
    if not user:
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    data = request.get_json() or {}
    run_id = data.get('run_id')
    
    if not run_id:
         return jsonify({"success": False, "error": "Missing run_id"}), 400

    # Basic ownership check before delete
    task = get_task(run_id)
    if task:
        if task.get("user") != user:
             return jsonify({"success": False, "error": "Unauthorized"}), 403
    
    # Delete the task
    try:
        delete_task_from_disk(run_id)
        return jsonify({"success": True})
    except Exception as e:
        print(f"Error removing task: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@socketio.on("cancel_task")
def handle_cancel_task(data=None):
    user = session.get("user")
    if not user:
        emit("status_update", {"error": "Unauthorized"}, room=request.sid)
        return
    
    run_id = data.get("run_id") if data else None
    if not run_id:
         # Fallback? No, unsafe.
         emit("status_update", {"error": "Missing run_id"}, room=request.sid)
         return

    task = get_task(run_id)
    if not task:
        emit("status_update", {"error": "No active task"}, room=request.sid)
        return

    if task.get("user") != user:
         emit("status_update", {"error": "Unauthorized"}, room=request.sid)
         return

    task["cancelled"] = True
    task["status"] = "cancelled"
    save_task(run_id, task)

    emit(
        "status_update",
        {"error": "Task cancelled by user"},
        room=f"task_{run_id}"
    )

@socketio.on("start_rank_check")
def handle_rank_check(data=None):
    print(f"DEBUG: handle_rank_check received for run_id {data.get('run_id') if data else 'None'}")
    user = session.get("user")
    if not user:
        emit("status_update", {"error": "Session expired. Login again."}, room=request.sid)
        return

    run_id = data.get("run_id") if data else None
    
    # Locked execution to prevent race condition (double task start)
    with TASK_LOCK:
        task = get_task(run_id)
        if not task:
            emit("status_update", {"error": "No task found."}, room=request.sid)
            return
        
        if task.get("user") != user:
            emit("status_update", {"error": "Unauthorized"}, room=request.sid)
            return

        if task.get("status") != "ready_to_start":
             # Allow re-join if started? Logic in UI handles this but backend should be safe.
             if task.get("status") == "started":
                 # Just ignore, it's running
                 return
             emit("status_update", {"error": "Task already running or finished."}, room=request.sid)
             return

        # Mark as started to prevent double-run
        task["status"] = "started"
        save_task(run_id, task)
    
    # Thread launch outside lock (or inside, but it's fast)
    keywords = task["keywords"]
    target_domain = task["target_domain"]
    max_pages = task["max_pages"]
    location = task.get("location", "Global")
    project_name = task.get("project_name", "No Project")
    sid = task.get("sid") # Legacy param, not used much if room used

    threading.Thread(
        target=run_rank_checker,
        args=(keywords, target_domain, max_pages, user, sid, project_name, location, run_id),
        daemon=True,
    ).start()
        

@app.route("/api/edit-project", methods=['POST'])
def api_edit_project():
    user = session.get("user")
    if not user:
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    data = request.get_json() or {}
    old_name = data.get('old_project_name')
    new_name = data.get('new_project_name')

    if not old_name or not new_name:
         return jsonify({"success": False, "error": "Missing parameters"}), 400

    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE runs 
                SET project_name = %s 
                WHERE project_name = %s AND user_email = %s
            """, (new_name, old_name, user))
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"success": True})
    except Exception as e:
        print(f"Error editing project: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    return jsonify({"success": False, "error": "DB Error"}), 500


@app.route("/api/delete-project", methods=['POST'])
def api_delete_project():
    user = session.get("user")
    if not user:
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    data = request.get_json() or {}
    project_name = data.get('project_name')

    if not project_name:
         return jsonify({"success": False, "error": "Missing parameters"}), 400

    try:
        conn = get_db_connection()
        if conn:
            # 1. Get all run IDs for this project
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT id FROM runs WHERE project_name = %s AND user_email = %s", (project_name, user))
            runs = cursor.fetchall()
            cursor.close()
            conn.close() # Close read connection

            # 2. Delete each task from disk
            for r in runs:
                try:
                    delete_task_from_disk(r['id'])
                except Exception as e:
                    print(f"Error deleting task {r['id']} files: {e}")
            
            # 3. Delete from DB (The individual delete_task_from_disk might delete by ID, 
            # but standard implementation often deletes from DB too. 
            # However, for efficiency, let's do a bulk DB delete if delete_task_from_disk doesn't enforce it OR 
            # just rely on loop if delete_task_from_disk does DB delete.
            # Checking `delete_task_from_disk` implementation usually reveals it does both. 
            # But to be safe and atomic-like for the project structure:
            
            conn2 = get_db_connection()
            if conn2:
                cursor2 = conn2.cursor()
                cursor2.execute("DELETE FROM runs WHERE project_name = %s AND user_email = %s", (project_name, user))
                conn2.commit()
                cursor2.close()
                conn2.close()
            
            return jsonify({"success": True})

    except Exception as e:
        print(f"Error deleting project: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    return jsonify({"success": False, "error": "DB Error"}), 500
@app.route("/api/delete-project-group", methods=['POST'])
def delete_project_group():
    user = session.get("user")
    if not user:
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    
    data = request.get_json()
    project_name = data.get('project_name')
    target_domain = data.get('target_domain')
    
    if not project_name or not target_domain:
         return jsonify({"success": False, "error": "Missing parameters"}), 400

    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("""
                DELETE FROM runs 
                WHERE user_email = %s AND project_name = %s AND target_domain = %s
            """, (user, project_name, target_domain))
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({"success": True})
    except Exception as e:
         print(f"Error deleting project group: {e}")
         return jsonify({"success": False, "error": str(e)}), 500

    return jsonify({"success": False, "error": "Database error"}), 500


@app.route("/api/external/start_task", methods=["POST"])
def api_external_start_task():
    try:
        # Support both JSON and Form Data
        data = request.get_json(silent=True)
        if not data:
            data = request.form
        
        # Map parameters
        # "priority" => 1 (Ignored)
        # "site" => target_domain
        # "se_id" => (Ignored)
        # "loc_id" => location
        # "max_crawl_pages" => max_pages
        # "key" => keyword
        
        target_domain = data.get("site")
        location = data.get("loc_id", "Global")
        max_pages = data.get("max_crawl_pages")
        keyword = data.get("key")
        
        if not target_domain:
            return jsonify({"success": False, "error": "Missing 'site' (target_domain)"}), 400
        if not keyword:
            return jsonify({"success": False, "error": "Missing 'key' (keyword)"}), 400

        # DATA SANITIZATION:
        # The user might send a full URL with fragments/queries (e.g. ".../#%20")
        # We should strip these to avoid breaking the substring match in find_keyword_rank.
        try:
            from urllib.parse import urlparse, urlunparse
            parsed = urlparse(target_domain)
            # Reconstruct without fragment or query, keep path
            # If scheme is missing (e.g. "example.com"), urlparse might put it in path.
            # But usually input has http/https if it's a URL.
            # If it's just "example.com", parsed.scheme is '', path is 'example.com'.
            
            clean_parts = list(parsed)
            clean_parts[4] = '' # query
            clean_parts[5] = '' # fragment
            target_domain = urlunparse(clean_parts)
            
            # If the result ends with /# or similar due to empty fragment? No, urlunparse handles it.
            # But if original was just "example.com", result is "example.com".
            # If "https://example.com/foo#bar", result is "https://example.com/foo".
        except Exception as e:
            print(f"Error sanitizing URL {target_domain}: {e}")
            # Fallback to original
            pass

        # Default max_pages
        try:
            max_pages = int(max_pages) if max_pages else 10
        except:
            max_pages = 10

        # Use a designated user for external 
        # tasks if not provided
        # This ensures these tasks are tracked but distinguishable
        user_email = data.get("user_email", "external_api@rankplex.com")
        
        # Convert single keyword to list as expected by internal logic
        keywords = [keyword]

        # Project Name logic
        project_name = data.get("project_name", "External API")
        
        # Start Task
        run_id = start_task_internal(
            keywords=keywords,
            target_domain=target_domain,
            location=location,
            project_name=project_name,
            max_pages=max_pages,
            user_email=user_email
        )
        
        return jsonify({
            "success": True, 
            "run_id": run_id, 
            "message": "Task started successfully via External API"
        })

    except Exception as e:
        print(f"External API Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=8000)
    args = parser.parse_args()

    init_db()

    socketio.run(
        app,
        host="0.0.0.0",
        port=args.port,
        debug=False,
        allow_unsafe_werkzeug=True,
    )
