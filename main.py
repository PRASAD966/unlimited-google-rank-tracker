from flask import Flask, render_template_string, request, redirect, url_for, session
from flask_socketio import SocketIO, emit
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
import pandas as pd
import time, os, json, tempfile, shutil, uuid, datetime, threading
from openpyxl import load_workbook
from openpyxl.styles import Font
import secrets, smtplib
from email.message import EmailMessage

# === CONFIG ===
BRAVE_PATH = r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"
EXTENSION_PATH = r"C:\Users\prasa\Downloads\key_word_rank\static\raptor_unpacked"
USERS_FILE = "users.json"
TOKENS_FILE = "tokens.json"
STATIC_FOLDER = "static"

# --- Email config (Gmail App Password recommended) ---
EMAIL_ADDRESS = "praveenniceinteractive@gmail.com"
EMAIL_PASSWORD = "tzvhtffluaxyeroq"

app = Flask(__name__, static_folder=STATIC_FOLDER)
app.secret_key = "super_secret_key_12345"
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="threading")

if not os.path.exists(STATIC_FOLDER):
    os.makedirs(STATIC_FOLDER)

TEMP_TASK_DATA = {}
LOCK = threading.Lock()

# === USER HELPERS ===
def load_users():
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, "r") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except:
            return {}
    return {}

def save_users(users):
    with open(USERS_FILE, "w") as f:
        json.dump(users, f, indent=4)

def load_tokens():
    if os.path.exists(TOKENS_FILE):
        try:
            with open(TOKENS_FILE, "r") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except:
            return {}
    return {}

def save_tokens(tokens):
    with open(TOKENS_FILE, "w") as f:
        json.dump(tokens, f, indent=4)

# === EMAIL SENDER ===
def send_reset_email(recipient, token):
    msg = EmailMessage()
    msg["Subject"] = "Password Reset Request"
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = recipient
    link = f"http://127.0.0.1:8000/reset-password/{token}"
    msg.set_content(f"Click the link below to reset your password:\n\n{link}\n\nIf you didn't request this, ignore this email.")

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(msg)

# === CAPTCHA CHECK ===
def wait_for_captcha_to_clear(driver):
    while True:
        try:
            page = driver.page_source.lower()
            if any(word in page for word in [
                "captcha", "i'm not a robot", "unusual traffic",
                "vaptcha", "recaptcha", "verify you are human", "solve the challenge"
            ]):
                time.sleep(10)
            else:
                return True
        except:
            time.sleep(10)

# === RANK FINDER ===
def find_keyword_rank(driver, keyword, target_domain, max_pages=10):
    rank, page_found = None, None
    domain_only = target_domain.replace("https://", "").replace("http://", "").replace("www.", "").strip("/")

    for page in range(max_pages):
        start = page * 10
        driver.get(f"https://www.google.co.in/search?q={keyword}&hl=en&gl=in&start={start}")
        wait_for_captcha_to_clear(driver)
        page_source = driver.page_source.lower()
        if "did not match any documents" in page_source:
            return None, None

        results = driver.find_elements(By.XPATH, "//a[h3]")
        organic_results = [r for r in results if r.get_attribute("href") and "google.com" not in r.get_attribute("href")]
        for index, result in enumerate(organic_results, start=1 + page * 10):
            href = result.get_attribute("href")
            if domain_only in href:
                return index, page + 1
    return None, None

# === WRITE EXCEL ===
def write_results_with_layout(file_path, target_domain, df):
    df_to_write = df[['Keyword', 'Page', 'Rank']].copy()
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df_to_write.to_excel(writer, index=False, startrow=5, header=False)
    wb = load_workbook(file_path)
    ws = wb.active
    ws["A1"] = "Target Domain"
    ws["A2"] = "Search Engine"
    ws["A3"] = "Date"
    ws["B1"] = target_domain
    ws["B2"] = "Google"
    ws["B3"] = datetime.datetime.now().strftime("%d-%m-%Y %H:%M")
    ws["A5"], ws["B5"], ws["C5"] = "Keyword", "Page", "Rank"
    bold = Font(bold=True)
    for cell in ["A1","A2","A3","A5","B5","C5"]:
        ws[cell].font = bold
    wb.save(file_path)

# === DRIVER CREATION HELPER ===
def create_uc_driver(options, sid=None, binary_path=None):
    """
    Create an undetected_chromedriver.Chrome instance that matches the installed browser major version.
    Emits socket status messages if sid provided.
    Returns driver or raises Exception.
    """
    try:
        # If user uses Brave (binary_path) we set options.binary_location already in caller.
        # Attempt to detect installed chrome version using uc utility:
        try:
            chrome_v = uc.find_chrome_v()  # returns e.g. '141.0.7390.122'
            chrome_major = int(str(chrome_v).split('.')[0])
        except Exception as e:
            # If detection fails, fallback to None -> undetected-chromedriver will pick latest compatible driver
            chrome_major = None

        if sid:
            socketio.emit("status_update", {"msg": "Launching auto-matching driver..."}, room=sid)

        driver = uc.Chrome(options=options, version_main=141)


        if sid:
            socketio.emit("status_update", {"msg": "Driver started successfully."}, room=sid)
        return driver

    except Exception as e:
        # Provide informative message including guidance about chrome/chromedriver mismatch
        err = str(e)
        guidance = ("Session couldn't be created. This often happens when Chrome/Brave and ChromeDriver versions don't match. "
                    "Ensure your browser is up-to-date or let the script auto-download the matching driver. "
                    "Full error: " + err)
        if sid:
            socketio.emit("status_update", {"error": guidance}, room=sid)
        # Re-raise to allow caller to handle cleanup
        raise

# === BACKGROUND RANK CHECK ===
def run_rank_checker(keywords, target_domain, max_pages, user_email, sid):
    temp_profile = tempfile.mkdtemp(prefix=f"brave_profile_{user_email}_")
    options = uc.ChromeOptions()
    options.binary_location = BRAVE_PATH
    options.add_argument(f"--user-data-dir={temp_profile}")
    # Visible browser mode (A): start maximized
    options.add_argument("--start-maximized")

    if os.path.exists(EXTENSION_PATH):
        options.add_argument(f"--load-extension={EXTENSION_PATH}")

    data_rows = []
    driver = None
    filename = None
    try:
        # Create driver (auto-matching ChromeDriver to installed Chrome/Brave)
        try:
            driver = create_uc_driver(options, sid=sid, binary_path=BRAVE_PATH)
        except Exception as e:
            # create_uc_driver already emitted error; ensure temp cleanup and return
            return

        total = len(keywords)
        last_filename = None
        for i, k in enumerate(keywords, start=1):
            if "near me" in k.lower():
                data_rows.append({"Keyword": k, "Page": "-", "Rank": "Skipped"})
            else:
                progress = int((i / total) * 100)
                socketio.emit("status_update", {"keyword": k, "progress": progress}, room=sid)
                try:
                    rank, page = find_keyword_rank(driver, k, target_domain, max_pages)
                    data_rows.append({"Keyword": k, "Page": page or "-", "Rank": rank or "Not found"})
                except Exception as inner_e:
                    # If any single keyword run errors (e.g., captcha/unexpected), log and continue
                    data_rows.append({"Keyword": k, "Page": "-", "Rank": f"Error: {inner_e}"})
                    socketio.emit("status_update", {"msg": f"Error searching '{k}': {inner_e}"}, room=sid)

            # save in batches
            if i % 10 == 0 or i == total:
                domain_safe = target_domain.replace("https://", "").replace("http://", "").replace("/", "_")
                filename = f"{domain_safe}_{'final' if i == total else f'batch_{(i-1)//10+1}'}.xlsx"
                file_path = os.path.join(STATIC_FOLDER, filename)
                df = pd.DataFrame(data_rows, columns=["Keyword", "Page", "Rank"])
                write_results_with_layout(file_path, target_domain, df)
                socketio.emit("status_update", {"batch_done": True, "filename": filename, "progress": int((i / total) * 100)}, room=sid)

        # Completed
        if filename:
            socketio.emit("status_update", {"done": True, "filename": filename, "progress": 100}, room=sid)
        else:
            socketio.emit("status_update", {"done": True, "filename": None, "progress": 100}, room=sid)

    except Exception as e:
        # Top-level exception — report to UI
        socketio.emit("status_update", {"error": f"Unexpected error: {e}"}, room=sid)
    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass
        shutil.rmtree(temp_profile, ignore_errors=True)

# === ROUTES ===
@app.route("/")
def home():
    if "user" in session:
        return redirect(url_for("dashboard"))
    return render_template_string("""
<html>
<head>
<title>Login</title>
<style>
body {margin:0;font-family:Poppins,sans-serif;background:#0a0a0f;display:flex;justify-content:center;align-items:center;height:100vh;}
.container{width:380px;background:linear-gradient(135deg,#2b0036,#4b0082);padding:40px;border-radius:20px;box-shadow:0 0 20px #7f00ff;color:white;}
h2{text-align:center;margin-bottom:20px;}
input{width:100%;padding:10px;margin:10px 0;border:none;border-radius:5px;}
button{width:100%;padding:10px;border:none;border-radius:5px;background:#9b5de5;color:white;font-weight:bold;cursor:pointer;}
p{text-align:center;margin-top:10px;}
a{color:#c77dff;text-decoration:none;}
</style>
</head>
<body>
                                  <!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Login</title>
  <link rel="stylesheet" href="{{ url_for('static', filename='style.css') }}">
  <style>
    body {
      margin: 0;
      height: 100vh;
      background: radial-gradient(circle at top left, #0a001a, #000);
      display: flex;
      flex-direction: column;
      align-items: center;
      justify-content: center;
      font-family: Arial, sans-serif;
      color: #fff;
    }

    .title-box {
      text-align: center;
      margin-bottom: 40px;
      color: #fff;
    }

    .title-box h1 {
      font-size: 2.2rem;
      color: #bb86fc;
      text-shadow: 0 0 20px rgba(187,134,252,0.8);
    }

    .login-container {
      background: linear-gradient(145deg, #2b0059, #6a00ff);
      padding: 40px;
      border-radius: 20px;
      box-shadow: 0 0 30px rgba(138,43,226,0.7);
      text-align: center;
      width: 350px;
    }

    .login-container h2 {
      margin-bottom: 20px;
      color: white;
    }

    input[type="email"], input[type="password"] {
      width: 100%;
      padding: 10px;
      margin: 10px 0;
      border: none;
      border-radius: 5px;
      outline: none;
    }

    button {
      width: 100%;
      background: #b266ff;
      color: white;
      border: none;
      padding: 10px;
      border-radius: 5px;
      font-weight: bold;
      cursor: pointer;
      transition: 0.3s;
    }

    button:hover {
      background: #a64dff;
    }

    p {
      margin-top: 15px;
      color: #ddd;
    }

    a {
      color: #bb86fc;
      text-decoration: none;
    }

    a:hover {
      text-decoration: underline;
    }
  </style>
</head>
<body>
  <div class="title-box">
    <h1>Nice Digitals Rank Checker</h1>
  </div>
<div class="container">
<h2>Login</h2>
<form method="post" action="{{ url_for('login') }}">
<input type="text" name="email" placeholder="Email" required>
<input type="password" name="password" placeholder="Password" required>
<button type="submit">Login</button>
</form>
<p><a href="{{ url_for('forgot_password_page') }}">Forgot Password?</a></p>
<p>Don't have an account? <a href="{{ url_for('signup_page') }}">Sign Up</a></p>
</div>
</body></html>
""")

@app.route("/signup")
def signup_page():
    return render_template_string("""
<html>
<head><title>Signup</title>
<style>
body{margin:0;font-family:Poppins,sans-serif;background:#0a0a0f;display:flex;justify-content:center;align-items:center;height:100vh;}
.container{width:380px;background:linear-gradient(135deg,#2b0036,#4b0082);padding:40px;border-radius:20px;box-shadow:0 0 20px #7f00ff;color:white;}
h2{text-align:center;margin-bottom:20px;}
input{width:100%;padding:10px;margin:10px 0;border:none;border-radius:5px;}
button{width:100%;padding:10px;border:none;border-radius:5px;background:#9b5de5;color:white;font-weight:bold;cursor:pointer;}
p{text-align:center;margin-top:10px;}
a{color:#c77dff;text-decoration:none;}
</style></head>
<body>
<div class="container">
<h2>Sign Up</h2>
<form method="post" action="{{ url_for('signup') }}">
<input type="text" name="email" placeholder="Email" required>
<input type="password" name="password" placeholder="Password" required>
<button type="submit">Register</button>
</form>
<p>Already have an account? <a href="{{ url_for('home') }}">Login</a></p>
</div>
</body></html>
""")

@app.route("/signup", methods=["POST"])
def signup():
    email, password = request.form.get("email"), request.form.get("password")
    users = load_users()
    if email in users:
        return "<h3 style='text-align:center;color:red;'>User exists. <a href='/'>Login</a></h3>"
    users[email] = password
    save_users(users)
    return "<h3 style='text-align:center;color:green;'>Registered. <a href='/'>Login</a></h3>"

@app.route("/forgot-password")
def forgot_password_page():
    return render_template_string("""
    <html><head><title>Forgot Password</title></head>
    <body style="margin:0;font-family:Poppins,sans-serif;background:#0a0a0f;display:flex;justify-content:center;align-items:center;height:100vh;">
    <div style="width:380px;background:linear-gradient(135deg,#2b0036,#4b0082);padding:40px;border-radius:20px;box-shadow:0 0 20px #7f00ff;color:white;">
    <h2>Forgot Password</h2>
    <form method="post" action="{{ url_for('forgot_password') }}">
        <input type="text" name="email" placeholder="Enter your registered email" required style="width:100%;padding:10px;margin:10px 0;border:none;border-radius:5px;">
        <button type="submit" style="width:100%;padding:10px;background:#9b5de5;color:white;border:none;border-radius:5px;">Send Reset Link</button>
    </form>
    <p style="text-align:center;"><a href="/">Back to Login</a></p>
    </div></body></html>
    """)

@app.route("/forgot-password", methods=["POST"])
def forgot_password():
    email = request.form.get("email")
    users = load_users()
    if email not in users:
        return "<h3 style='text-align:center;color:red;'>Email not found. <a href='/forgot-password'>Try again</a></h3>"
    token = secrets.token_urlsafe(16)
    tokens = load_tokens()
    tokens[token] = email
    save_tokens(tokens)
    try:
        send_reset_email(email, token)
    except Exception as e:
        return f"<h3 style='text-align:center;color:red;'>Failed to send email: {e}</h3>"
    return "<h3 style='text-align:center;color:green;'>Reset link sent to your email.</h3>"

@app.route("/reset-password/<token>")
def reset_password_page(token):
    tokens = load_tokens()
    if token not in tokens:
        return "<h3 style='text-align:center;color:red;'>Invalid or expired token.</h3>"
    return render_template_string(f"""
    <html><head><title>Reset Password</title></head>
    <body style="background:#0a0a0f;display:flex;justify-content:center;align-items:center;height:100vh;font-family:Poppins,sans-serif;color:white;">
    <div style="width:380px;background:linear-gradient(135deg,#2b0036,#4b0082);padding:40px;border-radius:20px;box-shadow:0 0 20px #7f00ff;">
    <h2>Reset Password</h2>
    <form method="post" action="/reset-password/{token}">
        <input type="password" name="password" placeholder="New Password" required style="width:100%;padding:10px;margin:10px 0;border:none;border-radius:5px;">
        <button type="submit" style="width:100%;padding:10px;background:#9b5de5;color:white;border:none;border-radius:5px;">Update Password</button>
    </form>
    </div></body></html>
    """)

@app.route("/reset-password/<token>", methods=["POST"])
def reset_password(token):
    tokens = load_tokens()
    if token not in tokens:
        return "<h3 style='text-align:center;color:red;'>Invalid or expired token.</h3>"
    email = tokens.pop(token)
    new_password = request.form.get("password")
    users = load_users()
    users[email] = new_password
    save_users(users)
    save_tokens(tokens)
    return "<h3 style='text-align:center;color:green;'>Password updated successfully. <a href='/'>Login</a></h3>"

@app.route("/login", methods=["POST"])
def login():
    email, password = request.form.get("email"), request.form.get("password")
    users = load_users()
    if email not in users or users[email] != password:
        return "<h3 style='text-align:center;color:red;'>Invalid credentials. <a href='/'>Try again</a></h3>"
    session["user"] = email
    return redirect(url_for("dashboard"))

@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("home"))

@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():
    if "user" not in session:
        return redirect(url_for("home"))
    return render_template_string("""
<html>
<head><title>Dashboard</title>
<style>
body{margin:0;font-family:Poppins,sans-serif;background:#0a0a0f;color:white;}
.sidebar{position:fixed;left:0;top:0;width:220px;height:100%;background:#1a001f;padding:20px;}
.sidebar h2{color:#c77dff;text-align:center;}
.sidebar a{display:block;color:white;padding:10px;margin:10px 0;text-decoration:none;border-radius:5px;}
.sidebar a:hover{background:#4b0082;}
.main{margin-left:240px;padding:30px;}
.container{background:#1a001f;padding:30px;border-radius:15px;box-shadow:0 0 15px #7f00ff;width:80%;margin:auto;}
input,button{padding:10px;border:none;border-radius:5px;margin:5px 0;}
button{background:#9b5de5;color:white;font-weight:bold;cursor:pointer;}
</style>
</head>
<body>
<div class="sidebar">
<h2>Dashboard</h2>
<p>{{ session['user'] }}</p>
<a href="/dashboard">🏠 Home</a>
<a href="/logout">🚪 Logout</a>
</div>
<div class="main">
<div class="container">
<h2>🔍 Keyword Rank Checker</h2>
<form method='POST' action='/start-task' enctype='multipart/form-data'>
<label>Upload Keyword File (.txt):</label><br>
<input type='file' name='file'><br><br>
<label>Or Enter Keywords:</label><br>
<div id="tagBox" style="width:80%;border:2px solid #999;padding:5px;min-height:45px;display:flex;flex-wrap:wrap;cursor:text;border-radius:5px;background:#fff;text-align:left;">
<input type="text" id="tagInput" style="border:none;outline:none;flex:1;padding:8px;font-size:15px;">
</div>
<input type="hidden" name="keywords" id="keywordsField"><br>
<label>Target Domain:</label><br>
<input type='text' name='target_domain' placeholder='https://example.com' style='width:80%;'><br><br>
<label>Max Pages:</label>
<input type='number' name='max_pages' value='10' min='1' style='width:60px;'><br><br>
<button type='submit'>🚀 Start Rank Check</button>
</form>
</div></div>
<script>
let tags = [];
document.getElementById('tagBox').addEventListener('click', ()=>document.getElementById('tagInput').focus());
document.getElementById('tagInput').addEventListener('keydown', function(e){
if(e.key==='Enter'){e.preventDefault();let v=this.value.trim();if(v!==""&&!tags.includes(v)){tags.push(v);addTag(v);}this.value="";}
});
function addTag(t){let box=document.getElementById('tagBox');let tag=document.createElement('span');
tag.textContent=t;tag.style="background:#9b5de5;color:#fff;padding:6px 10px;margin:3px;border-radius:4px;font-size:14px;";
let close=document.createElement('span');close.textContent=" ×";close.style="cursor:pointer;margin-left:5px;";
close.onclick=function(){tag.remove();tags=tags.filter(x=>x!==t);updateHidden();};tag.appendChild(close);
box.insertBefore(tag,document.getElementById('tagInput'));updateHidden();}
function updateHidden(){document.getElementById('keywordsField').value=tags.join(",");}
</script>
</body></html>
""")

@app.route("/start-task", methods=["POST"])
def start_task():
    user = session.get("user")
    if not user:
        return redirect(url_for("home"))
    keywords = []
    file = request.files.get("file")
    if file:
        content = file.read().decode()
        keywords = [x.strip() for x in content.splitlines() if x.strip()]
    kw_manual = request.form.get("keywords")
    if kw_manual:
        keywords += [x.strip() for x in kw_manual.split(",") if x.strip()]
    if not keywords:
        return "<h3>No keywords provided!</h3>"
    target_domain = request.form.get("target_domain")
    max_pages = int(request.form.get("max_pages", 5))
    with LOCK:
        TEMP_TASK_DATA[user] = {"keywords": keywords, "target_domain": target_domain, "max_pages": max_pages}
    return redirect(url_for("progress_page"))

@app.route("/progress")
def progress_page():
    user = session.get("user")
    if not user or user not in TEMP_TASK_DATA:
        return redirect(url_for("dashboard"))
    task_data = TEMP_TASK_DATA[user]
    return render_template_string("""
<html><head><title>Progress</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/socket.io/4.0.1/socket.io.js"></script>
<style>
body{background:#0a0a0f;color:white;font-family:Poppins;text-align:center;}
.container{background:#1a001f;padding:20px;border-radius:15px;box-shadow:0 0 15px #7f00ff;width:600px;margin:80px auto;}
#progress-bar{width:100%;background:#333;border-radius:5px;}
#fill{width:0%;height:30px;background:#9b5de5;color:white;border-radius:5px;line-height:30px;transition:width .3s;}
</style></head>
<body>
<div class="container">
<h3>Keyword Rank Check in Progress...</h3>
<p>Target: {{ target_domain }}</p>
<div id="progress-bar"><div id="fill">0%</div></div>
<p id="msg">Starting...</p>
<div id="links"></div>
</div>
<script>
var s=io();
s.on('connect',()=>s.emit('start_rank_check',{}));
s.on('status_update',m=>{
let f=document.getElementById('fill'),msg=document.getElementById('msg'),l=document.getElementById('links');
let p=m.progress||0;f.style.width=p+'%';f.textContent=p+'%';
if(m.error){msg.textContent=m.error;f.style.background='red';return;}
if(m.msg){msg.textContent=m.msg;}
if(m.batch_done&&m.filename){l.innerHTML+='<div><a href="/static/'+m.filename+'" style="color:#c77dff;">Download '+m.filename+'</a></div>';}
if(m.done&&m.filename){msg.textContent='Completed';f.style.background='green';l.innerHTML+='<div><a href="/static/'+m.filename+'" style="color:#c77dff;">Download Final</a></div>';}
});
</script></body></html>
""", target_domain=task_data["target_domain"])

@socketio.on("start_rank_check")
def handle_rank_check(data=None):
    user = session.get("user")
    if not user:
        emit("status_update", {"error": "Not logged in"})
        return
    with LOCK:
        task_data = TEMP_TASK_DATA.pop(user, None)
    if not task_data:
        emit("status_update", {"error": "No task data"})
        return
    # Run rank checker (synchronous within this thread) — create_uc_driver will emit driver start/fail messages
    run_rank_checker(task_data["keywords"], task_data["target_domain"], task_data["max_pages"], user, request.sid)

if __name__ == "__main__":
    socketio.run(app, host="0.0.0.0", port=8000, debug=True)
